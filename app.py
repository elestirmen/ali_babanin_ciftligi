"""
Ali Baba'nin Ciftligi - Ana Uygulama
Ogul kovanlari ve sabit arilik takip sistemi.
"""

import os
import csv
import io
import re
import secrets
import sqlite3
import tempfile
import uuid
import zipfile
from datetime import datetime

from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, session, abort, send_from_directory,
                   Response, send_file)
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.utils import secure_filename
from PIL import Image, ImageOps
from PIL.ExifTags import TAGS, GPSTAGS
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import qrcode

# ---------------------------------------------------------------------------
# Uygulama konfigurasyonu
# ---------------------------------------------------------------------------
app = Flask(__name__)
if os.environ.get('ALI_BABA_PROXY_FIX', '0') == '1':
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.environ.get('ALI_BABA_DB_PATH', os.path.join(BASE_DIR, 'ali_baba.db'))
UPLOAD_FOLDER = os.environ.get('ALI_BABA_UPLOAD_FOLDER',
                               os.path.join(BASE_DIR, 'uploads'))
QR_FOLDER = os.environ.get('ALI_BABA_QR_FOLDER',
                           os.path.join(BASE_DIR, 'qrcodes'))
PUBLIC_UPLOAD_FOLDER = os.environ.get('ALI_BABA_PUBLIC_UPLOAD_FOLDER',
                                      os.path.join(BASE_DIR, 'public_uploads'))
BACKUP_FOLDER = os.environ.get('ALI_BABA_BACKUP_FOLDER',
                               os.path.join(os.path.dirname(DB_PATH), 'backups'))
PUBLIC_URL = os.environ.get('ALI_BABA_PUBLIC_URL', '').rstrip('/')
APP_PASSWORD = os.environ.get('ALI_BABA_PASSWORD', 'alibaba')
if os.environ.get('ALI_BABA_REQUIRE_SECURE_PASSWORD', '0') == '1':
    insecure_passwords = {'alibaba', 'change-me', 'password', 'admin'}
    if APP_PASSWORD.strip().lower() in insecure_passwords:
        raise RuntimeError('ALI_BABA_PASSWORD guclu bir parola olarak ayarlanmalidir.')
ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png', 'webp'}

ASSET_VERSION = os.environ.get('ALI_BABA_ASSET_VERSION')
if not ASSET_VERSION:
    asset_version_paths = [
        __file__,
        os.path.join(BASE_DIR, 'static', 'css', 'style.css'),
        os.path.join(BASE_DIR, 'static', 'js', 'map.js'),
        os.path.join(BASE_DIR, 'static', 'js', 'location_picker.js'),
        os.path.join(BASE_DIR, 'static', 'js', 'public-guide.js'),
        os.path.join(BASE_DIR, 'static', 'js', 'pwa.js'),
        os.path.join(BASE_DIR, 'static', 'sw.js'),
    ]
    ASSET_VERSION = str(int(max(
        os.path.getmtime(path)
        for path in asset_version_paths
        if os.path.exists(path)
    )))

app.secret_key = os.environ.get('ALI_BABA_SECRET_KEY') or secrets.token_hex(32)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['QR_FOLDER'] = QR_FOLDER
app.config['PUBLIC_UPLOAD_FOLDER'] = PUBLIC_UPLOAD_FOLDER
app.config['BACKUP_FOLDER'] = BACKUP_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB

# Klasorleri olustur
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(QR_FOLDER, exist_ok=True)
os.makedirs(PUBLIC_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BACKUP_FOLDER, exist_ok=True)

# ---------------------------------------------------------------------------
# Durum secenekleri
# ---------------------------------------------------------------------------
SWARM_STATUSES = [
    'Boş', 'Arı hareketi var', 'Oğul girdi',
    'Ana arı kontrol edildi', 'Taşındı', 'İptal edildi'
]

SWARM_ARRIVAL_STATUSES = {
    'Arı hareketi var', 'Oğul girdi', 'Ana arı kontrol edildi'
}

FIXED_HIVE_STATUSES = [
    'Zayıf', 'Orta', 'Güçlü', 'Ana arı sorunu var',
    'Hastalık şüphesi var', 'Bal durumu iyi', 'Kontrol gerekli', 'Pasif'
]

INSPECTION_CHOICES = {
    'ari_yogunlugu': ['Zayıf', 'Orta', 'Güçlü'],
    'yavru_durumu': ['Yok', 'Zayıf', 'Normal', 'Çok iyi'],
    'bal_durumu': ['Az', 'Orta', 'İyi'],
    'polen_gelisi': ['Yok', 'Az', 'Normal', 'Yoğun'],
    'ana_ari_goruldu': ['Evet', 'Hayır'],
    'yumurta_var': ['Evet', 'Hayır'],
    'hastalik_belirtisi': ['Yok', 'Var', 'Şüpheli'],
    'saldirganlik': ['Düşük', 'Orta', 'Yüksek'],
    'besleme_yapildi': ['Evet', 'Hayır'],
    'ilaclama_yapildi': ['Evet', 'Hayır'],
}

GENERAL_MESSAGE_CATEGORIES = [
    'Bal hakkında soru', 'Ziyaret talebi', 'Arıcılık sorusu', 'İş birliği', 'Diğer'
]

HIVE_MESSAGE_CATEGORIES = [
    'Kovan devrilmiş / zarar görmüş',
    'Arılarla ilgili tehlike fark ettim',
    'Kovan yeriyle ilgili bilgi vermek istiyorum',
    'Bal / ürün hakkında bilgi almak istiyorum',
    'Diğer',
]

PUBLIC_MESSAGE_CATEGORIES = list(dict.fromkeys(
    HIVE_MESSAGE_CATEGORIES + GENERAL_MESSAGE_CATEGORIES
))

PUBLIC_MESSAGE_STATUSES = ['Yeni', 'Okundu', 'Yanıtlandı', 'Ziyaret planlandı']

PUBLIC_POST_CATEGORIES = ['Flora', 'Kovan Bakımı', 'Hasat', 'Ziyaret', 'Duyuru']

# ---------------------------------------------------------------------------
# Basit oturum ve CSRF korumasi
# ---------------------------------------------------------------------------

PUBLIC_ENDPOINTS = {
    'login',
    'static',
    'public_home',
    'public_message',
    'public_qr_fixed',
    'public_qr_swarm',
    'fixed_hive_detail',
    'swarm_detail',
    'public_journal',
    'public_honey_stories',
    'public_honey_story_detail',
    'offline',
    'manifest',
    'service_worker',
    'public_upload',
}


def is_safe_local_next(value):
    """Sadece uygulama icindeki goreli yonlendirmelere izin verir."""
    return bool(value) and value.startswith('/') and not value.startswith('//')


def csrf_token():
    """Oturum bazli CSRF token dondurur."""
    token = session.get('_csrf_token')
    if not token:
        token = secrets.token_urlsafe(32)
        session['_csrf_token'] = token
    return token


def build_public_url(endpoint, **values):
    """Public URL tanimliysa mutlak URL'leri o domain ile uretir."""
    path = url_for(endpoint, **values)
    if PUBLIC_URL:
        return f"{PUBLIC_URL}{path}"
    return request.host_url.rstrip('/') + path


def parse_map_focus(value):
    """Harita odak parametresini (swarm-1/fixed-2/apiary-3) ayrıştırır."""
    match = re.fullmatch(r'(swarm|fixed|apiary)-(\d+)', value or '')
    if not match:
        return None, None
    return match.group(1), int(match.group(2))


def normalize_media_path(value):
    """Veritabaninda saklanan medya yolunu tek bicime getirir."""
    return (value or '').replace('\\', '/').lstrip('/')


def media_url(value):
    """Medya yolunu uygun public veya auth kontrollu URL'e cevirir."""
    path = normalize_media_path(value)
    if path.startswith('uploads/'):
        return url_for('private_upload', filename=path.removeprefix('uploads/'))
    if path.startswith('qrcodes/'):
        return url_for('private_qrcode', filename=path.removeprefix('qrcodes/'))
    if path.startswith('public_uploads/'):
        return url_for('public_upload', filename=path.removeprefix('public_uploads/'))
    return url_for('static', filename=path)


def media_abs_path(value):
    """Medya yolunun diskteki karsiligini dondurur."""
    path = normalize_media_path(value)
    if path.startswith('uploads/'):
        return os.path.join(UPLOAD_FOLDER, path.removeprefix('uploads/'))
    if path.startswith('qrcodes/'):
        return os.path.join(QR_FOLDER, path.removeprefix('qrcodes/'))
    if path.startswith('public_uploads/'):
        return os.path.join(PUBLIC_UPLOAD_FOLDER, path.removeprefix('public_uploads/'))
    return os.path.join(BASE_DIR, 'static', path)


@app.before_request
def require_login_and_csrf():
    """Statik dosyalar ve login disinda tum sayfalari basit sifreyle korur."""
    if request.endpoint == 'static':
        filename = normalize_media_path((request.view_args or {}).get('filename'))
        if filename.startswith(('uploads/', 'qrcodes/')) and not session.get('authenticated'):
            abort(404)
        return None

    if request.endpoint in PUBLIC_ENDPOINTS or request.endpoint is None:
        return None

    if not session.get('authenticated'):
        next_url = request.full_path if request.query_string else request.path
        return redirect(url_for('login', next=next_url))

    if request.method in {'POST', 'PUT', 'PATCH', 'DELETE'}:
        expected = session.get('_csrf_token')
        supplied = request.form.get('csrf_token') or request.headers.get('X-CSRF-Token')
        if not expected or not supplied or not secrets.compare_digest(expected, supplied):
            abort(400, description='CSRF token gecersiz veya eksik.')

    return None

# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar: Veritabani
# ---------------------------------------------------------------------------

def get_db():
    """Veritabani baglantisi dondurur."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def query_db(query, args=(), one=False):
    """Veritabani sorgusu calistirir ve sonuclari dondurur."""
    conn = get_db()
    cursor = conn.execute(query, args)
    results = cursor.fetchall()
    conn.close()
    if one:
        return results[0] if results else None
    return results


def execute_db(query, args=()):
    """Veritabaninda INSERT/UPDATE/DELETE calistirir, lastrowid dondurur."""
    conn = get_db()
    cursor = conn.execute(query, args)
    conn.commit()
    lastrowid = cursor.lastrowid
    conn.close()
    return lastrowid


def ensure_runtime_schema():
    """Mevcut kurulumlarda yeni public mesaj kolonlarini ekler."""
    if not os.path.exists(DB_PATH):
        return

    conn = get_db()
    try:
        table_exists = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='public_messages'"
        ).fetchone()
        if not table_exists:
            return

        columns = {
            row[1]
            for row in conn.execute("PRAGMA table_info(public_messages)").fetchall()
        }
        if 'kaynak_turu' not in columns:
            conn.execute('ALTER TABLE public_messages ADD COLUMN kaynak_turu TEXT')
        if 'kaynak_id' not in columns:
            conn.execute('ALTER TABLE public_messages ADD COLUMN kaynak_id INTEGER')
        conn.execute('''
            CREATE INDEX IF NOT EXISTS idx_public_messages_source
            ON public_messages (kaynak_turu, kaynak_id)
        ''')
        conn.commit()
    finally:
        conn.close()


ensure_runtime_schema()


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar: Yedekleme ve disa aktarma
# ---------------------------------------------------------------------------

BACKUP_FILENAME_RE = re.compile(r'^ali_baba_backup_\d{8}_\d{6}\.zip$')


def google_maps_directions_url(lat, lng):
    """Koordinat varsa Google Maps yol tarifi URL'i dondurur."""
    if lat is None or lng is None:
        return ''
    return f'https://www.google.com/maps/dir/?api=1&destination={lat},{lng}'


def yes_no(value):
    """Boolean/int alanlari arayuz ve raporlar icin okunur metne cevirir."""
    return 'Evet' if int(value or 0) == 1 else 'Hayır'


def export_date_suffix():
    return datetime.now().strftime('%Y%m%d')


def backup_timestamp():
    return datetime.now().strftime('%Y%m%d_%H%M%S')


def backup_path_for(filename):
    if not BACKUP_FILENAME_RE.match(filename or ''):
        return None
    return os.path.join(BACKUP_FOLDER, filename)


def list_backup_files():
    """Yedek klasorundeki zip dosyalarini yeni tarihten eskiye listeler."""
    if not os.path.isdir(BACKUP_FOLDER):
        return []

    backups = []
    for filename in os.listdir(BACKUP_FOLDER):
        if not BACKUP_FILENAME_RE.match(filename):
            continue
        path = os.path.join(BACKUP_FOLDER, filename)
        if not os.path.isfile(path):
            continue
        stat = os.stat(path)
        backups.append({
            'filename': filename,
            'size': stat.st_size,
            'size_label': format_file_size(stat.st_size),
            'created_at': datetime.fromtimestamp(stat.st_mtime).strftime('%d.%m.%Y %H:%M'),
            'mtime': stat.st_mtime,
        })
    return sorted(backups, key=lambda item: item['mtime'], reverse=True)


def format_file_size(size):
    if size < 1024:
        return f'{size} B'
    if size < 1024 * 1024:
        return f'{size / 1024:.1f} KB'
    return f'{size / (1024 * 1024):.1f} MB'


def add_sqlite_database_to_zip(zip_file):
    """SQLite DB'yi tutarli bir anlik kopya olarak zip'e ekler."""
    if not os.path.exists(DB_PATH):
        return False

    fd, temp_path = tempfile.mkstemp(prefix='ali_baba_db_', suffix='.db')
    os.close(fd)
    try:
        source = sqlite3.connect(DB_PATH)
        try:
            target = sqlite3.connect(temp_path)
            try:
                source.backup(target)
            finally:
                target.close()
        finally:
            source.close()
        zip_file.write(temp_path, os.path.basename(DB_PATH))
        return True
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def add_folder_to_zip(zip_file, folder_path, archive_root):
    """Var olan klasoru zip'e ekler; eksik klasorler sessizce atlanir."""
    if not os.path.isdir(folder_path):
        return 0

    added = 0
    for root, _dirs, files in os.walk(folder_path):
        for filename in files:
            source = os.path.join(root, filename)
            if os.path.islink(source):
                continue
            relative = os.path.relpath(source, folder_path)
            zip_file.write(source, os.path.join(archive_root, relative))
            added += 1
    return added


def cleanup_old_backups(keep=10):
    for backup in list_backup_files()[keep:]:
        path = backup_path_for(backup['filename'])
        if path and os.path.exists(path):
            os.remove(path)


def create_backup_zip():
    """DB ve veri klasorlerinden zip yedegi olusturur."""
    os.makedirs(BACKUP_FOLDER, exist_ok=True)
    filename = f'ali_baba_backup_{backup_timestamp()}.zip'
    path = os.path.join(BACKUP_FOLDER, filename)
    included = []

    with zipfile.ZipFile(path, 'w', compression=zipfile.ZIP_DEFLATED) as zip_file:
        if add_sqlite_database_to_zip(zip_file):
            included.append(os.path.basename(DB_PATH))

        folder_specs = [
            (UPLOAD_FOLDER, 'uploads'),
            (QR_FOLDER, 'qrcodes'),
            (PUBLIC_UPLOAD_FOLDER, 'public_uploads'),
        ]
        for folder_path, archive_root in folder_specs:
            count = add_folder_to_zip(zip_file, folder_path, archive_root)
            if count:
                included.append(archive_root)

    if not included:
        os.remove(path)
        raise RuntimeError('Yedeklenecek veritabanı veya dosya klasörü bulunamadı.')

    cleanup_old_backups()
    return filename


def csv_download_response(filename, headers, rows):
    output = io.StringIO(newline='')
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerows(rows)
    data = output.getvalue().encode('utf-8-sig')
    response = Response(data, content_type='text/csv; charset=utf-8')
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def swarm_export_data():
    headers = [
        'ID', 'Kovan adı', 'Durum', 'Latitude', 'Longitude', 'Kurulum tarihi',
        'Son kontrol tarihi', 'Aktif', 'Ulaşım notu', 'Genel not',
        'Google Maps yol tarifi linki'
    ]
    rows = []
    for hive in query_db('SELECT * FROM swarm_hives ORDER BY id'):
        rows.append([
            hive['id'],
            hive['ad'],
            hive['durum'],
            hive['latitude'],
            hive['longitude'],
            hive['kurulum_tarihi'],
            hive['son_kontrol_tarihi'],
            yes_no(hive['aktif']),
            hive['ulasim_notu'],
            hive['genel_not'],
            google_maps_directions_url(hive['latitude'], hive['longitude']),
        ])
    return headers, rows


def fixed_hive_export_data():
    headers = [
        'ID', 'Arılık adı', 'Kovan no', 'Sıra no', 'Konum no', 'Ana arı yılı',
        'Kat sayısı', 'Çerçeve sayısı', 'Durum', 'Son kontrol tarihi',
        'Aktif', 'Genel not', 'Google Maps yol tarifi linki'
    ]
    rows = []
    hives = query_db('''
        SELECT fh.*, a.arilik_adi, a.latitude, a.longitude
        FROM fixed_hives fh
        JOIN apiaries a ON a.id = fh.arilik_id
        ORDER BY a.arilik_adi, fh.sira_no, fh.konum_no, fh.id
    ''')
    for hive in hives:
        rows.append([
            hive['id'],
            hive['arilik_adi'],
            hive['kovan_no'],
            hive['sira_no'],
            hive['konum_no'],
            hive['ana_ari_yili'],
            hive['kat_sayisi'],
            hive['cerceve_sayisi'],
            hive['durum'],
            hive['son_kontrol_tarihi'],
            yes_no(hive['aktif']),
            hive['genel_not'],
            google_maps_directions_url(hive['latitude'], hive['longitude']),
        ])
    return headers, rows


def inspection_export_data():
    headers = [
        'ID', 'Arılık adı', 'Kovan no', 'Kontrol tarihi', 'Arı yoğunluğu',
        'Yavru durumu', 'Bal durumu', 'Polen gelişi', 'Ana arı görüldü mü',
        'Yumurta var mı', 'Hastalık belirtisi', 'Saldırganlık',
        'Besleme yapıldı mı', 'İlaçlama yapıldı mı', 'Notlar'
    ]
    rows = []
    inspections = query_db('''
        SELECT i.*, fh.kovan_no, a.arilik_adi
        FROM inspections i
        JOIN fixed_hives fh ON fh.id = i.kovan_id
        JOIN apiaries a ON a.id = fh.arilik_id
        ORDER BY i.kontrol_tarihi DESC, i.id DESC
    ''')
    for inspection in inspections:
        rows.append([
            inspection['id'],
            inspection['arilik_adi'],
            inspection['kovan_no'],
            inspection['kontrol_tarihi'],
            inspection['ari_yogunlugu'],
            inspection['yavru_durumu'],
            inspection['bal_durumu'],
            inspection['polen_gelisi'],
            inspection['ana_ari_goruldu'],
            inspection['yumurta_var'],
            inspection['hastalik_belirtisi'],
            inspection['saldirganlik'],
            inspection['besleme_yapildi'],
            inspection['ilaclama_yapildi'],
            inspection['notlar'],
        ])
    return headers, rows


def apiary_export_data():
    headers = [
        'ID', 'Arılık adı', 'Latitude', 'Longitude', 'Açıklama',
        'Oluşturma tarihi', 'Güncelleme tarihi'
    ]
    rows = []
    for apiary in query_db('SELECT * FROM apiaries ORDER BY arilik_adi, id'):
        rows.append([
            apiary['id'],
            apiary['arilik_adi'],
            apiary['latitude'],
            apiary['longitude'],
            apiary['aciklama'],
            apiary['olusturma_tarihi'],
            apiary['guncelleme_tarihi'],
        ])
    return headers, rows


def apiary_summary_export_data():
    headers = [
        'Arılık ID', 'Arılık adı', 'Latitude', 'Longitude',
        'Toplam kovan sayısı', 'Aktif kovan sayısı', 'Pasif kovan sayısı',
        'Güçlü kovan sayısı', 'Orta kovan sayısı', 'Zayıf kovan sayısı',
        'Kontrol gerekli kovan sayısı', 'Hastalık şüphesi olan kovan sayısı',
        'En eski son kontrol tarihi'
    ]
    rows = []
    summaries = query_db('''
        SELECT
            a.id,
            a.arilik_adi,
            a.latitude,
            a.longitude,
            COUNT(fh.id) as toplam_kovan_sayisi,
            SUM(CASE WHEN fh.aktif = 1 THEN 1 ELSE 0 END) as aktif_kovan_sayisi,
            SUM(CASE WHEN fh.aktif = 0 THEN 1 ELSE 0 END) as pasif_kovan_sayisi,
            SUM(CASE WHEN fh.durum = 'Güçlü' THEN 1 ELSE 0 END) as guclu_kovan_sayisi,
            SUM(CASE WHEN fh.durum = 'Orta' THEN 1 ELSE 0 END) as orta_kovan_sayisi,
            SUM(CASE WHEN fh.durum = 'Zayıf' THEN 1 ELSE 0 END) as zayif_kovan_sayisi,
            SUM(CASE WHEN fh.durum = 'Kontrol gerekli' THEN 1 ELSE 0 END) as kontrol_gerekli_kovan_sayisi,
            SUM(CASE WHEN fh.durum = 'Hastalık şüphesi var' THEN 1 ELSE 0 END) as hastalik_supheli_kovan_sayisi,
            MIN(fh.son_kontrol_tarihi) as en_eski_son_kontrol_tarihi
        FROM apiaries a
        LEFT JOIN fixed_hives fh ON fh.arilik_id = a.id
        GROUP BY a.id
        ORDER BY a.arilik_adi, a.id
    ''')
    for summary in summaries:
        rows.append([
            summary['id'],
            summary['arilik_adi'],
            summary['latitude'],
            summary['longitude'],
            summary['toplam_kovan_sayisi'] or 0,
            summary['aktif_kovan_sayisi'] or 0,
            summary['pasif_kovan_sayisi'] or 0,
            summary['guclu_kovan_sayisi'] or 0,
            summary['orta_kovan_sayisi'] or 0,
            summary['zayif_kovan_sayisi'] or 0,
            summary['kontrol_gerekli_kovan_sayisi'] or 0,
            summary['hastalik_supheli_kovan_sayisi'] or 0,
            summary['en_eski_son_kontrol_tarihi'],
        ])
    return headers, rows


def add_maps_link_column(headers, rows):
    """Latitude/Longitude kolonlari olan Excel sayfalarina link kolonu ekler."""
    if 'Latitude' not in headers or 'Longitude' not in headers:
        return headers, rows

    lat_index = headers.index('Latitude')
    lng_index = headers.index('Longitude')
    extended_headers = list(headers) + ['Google Maps yol tarifi linki']
    extended_rows = []
    for row in rows:
        lat = row[lat_index]
        lng = row[lng_index]
        extended_rows.append(list(row) + [google_maps_directions_url(lat, lng)])
    return extended_headers, extended_rows


def apiary_excel_data():
    return add_maps_link_column(*apiary_export_data())


def apiary_summary_excel_data():
    return add_maps_link_column(*apiary_summary_export_data())


def style_google_maps_links(sheet, headers):
    """Google Maps URL hucrelerini Excel hyperlink olarak isaretler."""
    for column_index, header in enumerate(headers, start=1):
        if header != 'Google Maps yol tarifi linki':
            continue
        for row_index in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = 'Hyperlink'


def add_excel_table(sheet, table_name, headers):
    """Sheet araligini filtrelenebilir/siralanabilir Excel tablosuna cevirir."""
    if not headers:
        return
    last_column = get_column_letter(len(headers))
    ref = f'A1:{last_column}{sheet.max_row}'
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def populate_excel_sheet(sheet, headers, rows, table_name):
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(['' if value is None else value for value in row])
    style_google_maps_links(sheet, headers)
    add_excel_table(sheet, table_name, headers)

    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            max_length = max(max_length, len(str(cell.value or '')))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
    sheet.freeze_panes = 'A2'


def add_excel_sheet(workbook, title, headers, rows, table_name):
    sheet = workbook.create_sheet(title)
    populate_excel_sheet(sheet, headers, rows, table_name)


def excel_download_response():
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, (title, data_func) in enumerate([
        ('Oğul Kovanları', swarm_export_data),
        ('Arılıklar', apiary_excel_data),
        ('Sabit Kovanlar', fixed_hive_export_data),
        ('Kontrol Kayıtları', inspection_export_data),
        ('Arılık Özeti', apiary_summary_excel_data),
    ], start=1):
        headers, rows = data_func()
        add_excel_sheet(workbook, title, headers, rows, f'ExportTable{index}')

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'ali_baba_tum_veriler_{export_date_suffix()}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def single_sheet_excel_response(sheet_title, filename_prefix, data_func):
    headers, rows = data_func()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    populate_excel_sheet(sheet, headers, rows, 'ExportTable1')

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'{filename_prefix}_{export_date_suffix()}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def export_dataset_options():
    """Disa aktarma sayfasinda tablo halinde gosterilecek veri setleri."""
    return [
        {
            'title': 'Oğul Kovanları',
            'description': 'Oğul kovan konumları, durumları, notları ve yol tarifi linkleri.',
            'count': len(swarm_export_data()[1]),
            'csv_endpoint': 'admin_export_swarm_hives_csv',
            'xlsx_endpoint': 'admin_export_swarm_hives_xlsx',
        },
        {
            'title': 'Sabit Kovanlar',
            'description': 'Arılık adı, kroki bilgileri, durum, aktiflik ve yol tarifi linkleri.',
            'count': len(fixed_hive_export_data()[1]),
            'csv_endpoint': 'admin_export_fixed_hives_csv',
            'xlsx_endpoint': 'admin_export_fixed_hives_xlsx',
        },
        {
            'title': 'Kontrol Kayıtları',
            'description': 'Sabit kovan kontrol tarihleri, gözlemler ve bakım notları.',
            'count': len(inspection_export_data()[1]),
            'csv_endpoint': 'admin_export_inspections_csv',
            'xlsx_endpoint': 'admin_export_inspections_xlsx',
        },
        {
            'title': 'Arılık Özeti',
            'description': 'Arılık bazında toplamlar, durum kırılımları ve en eski kontrol tarihi.',
            'count': len(apiary_summary_export_data()[1]),
            'csv_endpoint': 'admin_export_apiary_summary_csv',
            'xlsx_endpoint': 'admin_export_apiary_summary_xlsx',
        },
    ]


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar: Koordinat dogrulama
# ---------------------------------------------------------------------------

def safe_float(value, min_val=None, max_val=None):
    """
    Kullanicidan gelen koordinat degerini guvenli float'a cevirir.
    - Virgulu noktaya cevirir.
    - Gecersiz sayida None dondurur.
    - min_val/max_val varsa aralik kontrolu yapar; disindaysa None dondurur.
    Donus: (float_deger veya None, hata_mesaji veya None)
    """
    if value is None:
        return None, None
    value = str(value).strip()
    if not value:
        return None, None

    # Virgulu noktaya cevir
    value = value.replace(',', '.')

    try:
        result = float(value)
    except (ValueError, TypeError):
        return None, f"'{value}' gecerli bir sayi degil."

    if min_val is not None and result < min_val:
        return None, f"Deger {min_val} ile {max_val} arasinda olmali (girilen: {result})."
    if max_val is not None and result > max_val:
        return None, f"Deger {min_val} ile {max_val} arasinda olmali (girilen: {result})."

    return result, None


def parse_coordinates(lat_str, lng_str):
    """
    Latitude ve longitude string degerlerini dogrulayarak float ciftine cevirir.
    Donus: (lat, lng, hata_listesi)
    """
    errors = []

    lat_val, lat_err = safe_float(lat_str, min_val=-90, max_val=90)
    if lat_err:
        errors.append(f"Enlem: {lat_err}")

    lng_val, lng_err = safe_float(lng_str, min_val=-180, max_val=180)
    if lng_err:
        errors.append(f"Boylam: {lng_err}")

    return lat_val, lng_val, errors


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar: Form dogrulama
# ---------------------------------------------------------------------------

def validate_choice(value, choices, field_label):
    """Formdan gelen secim izinli listede mi kontrol eder."""
    if value in choices:
        return value, None
    return None, f"{field_label}: Gecersiz secim."


def validate_date(value, field_label, required=False):
    """YYYY-MM-DD tarihini dogrular."""
    value = (value or '').strip()
    if not value:
        if required:
            return None, f"{field_label}: Tarih zorunludur."
        return None, None
    try:
        datetime.strptime(value, '%Y-%m-%d')
    except ValueError:
        return None, f"{field_label}: Gecersiz tarih."
    return value, None


def parse_int_field(value, field_label, default=None, min_val=None, max_val=None):
    """Sayi alanlarini int'e cevirir ve sinirlarini kontrol eder."""
    value = (value or '').strip()
    if not value:
        return default, None
    try:
        parsed = int(value)
    except (ValueError, TypeError):
        return None, f"{field_label}: Gecerli bir sayi giriniz."

    if min_val is not None and parsed < min_val:
        return None, f"{field_label}: En az {min_val} olmali."
    if max_val is not None and parsed > max_val:
        return None, f"{field_label}: En fazla {max_val} olmali."
    return parsed, None


def flash_errors(errors):
    """Bos olmayan hata mesajlarini flash ile gosterir."""
    for err in errors:
        if err:
            flash(err, 'error')


def redirect_to_next(default_endpoint, **values):
    """Formdan gelen guvenli next varsa oraya, yoksa verilen endpoint'e doner."""
    next_url = request.form.get('next')
    if is_safe_local_next(next_url):
        return redirect(next_url)
    return redirect(url_for(default_endpoint, **values))


def delete_confirmation_valid():
    """Silme islemlerinde kullanicinin SIL onayini yazdigini kontrol eder."""
    confirmation = request.form.get('confirm_text', '').strip().upper()
    return confirmation in {'SIL', 'SİL'}


def create_swarm_inspection_record(swarm_hive_id, kontrol_tarihi, durum,
                                   ari_gelis_tarihi=None, notlar=None, fotograf_yolu=None):
    """Ogul kovani kontrol kaydini ortak sekilde olusturur."""
    return execute_db('''
        INSERT INTO swarm_inspections
        (swarm_hive_id, kontrol_tarihi, durum, ari_gelis_tarihi, notlar, fotograf_yolu)
        VALUES (?, ?, ?, ?, ?, ?)
    ''', (swarm_hive_id, kontrol_tarihi, durum, ari_gelis_tarihi,
          notlar or None, fotograf_yolu))


def fixed_hive_conflict(apiary_id, kovan_no, sira_no, konum_no, exclude_id=None):
    """Ayni arilik icinde kovan no veya kroki konumu cakismasini bulur."""
    params = [apiary_id, kovan_no]
    exclude_sql = ''
    if exclude_id is not None:
        exclude_sql = ' AND id != ?'
        params.append(exclude_id)

    existing_no = query_db(f'''
        SELECT id FROM fixed_hives
        WHERE arilik_id = ? AND lower(kovan_no) = lower(?){exclude_sql}
        LIMIT 1
    ''', params, one=True)
    if existing_no:
        return 'Bu arilikta ayni kovan numarasi zaten var.'

    params = [apiary_id, sira_no, konum_no]
    if exclude_id is not None:
        params.append(exclude_id)
    existing_position = query_db(f'''
        SELECT id FROM fixed_hives
        WHERE arilik_id = ? AND sira_no = ? AND konum_no = ?{exclude_sql}
        LIMIT 1
    ''', params, one=True)
    if existing_position:
        return 'Bu arilikta ayni sira/konumda baska bir kovan var.'

    return None


# ---------------------------------------------------------------------------
# Yardimci fonksiyonlar: Dosya ve EXIF
# ---------------------------------------------------------------------------

def allowed_file(filename):
    """Dosya uzantisi kabul edilen formatlar arasinda mi kontrol eder."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def verify_image(filepath):
    """
    Dosyanin gercekten bir goruntu olup olmadigini Pillow ile dogrular.
    Gecerliyse True, degilse False dondurur.
    """
    try:
        with Image.open(filepath) as img:
            img.verify()
        return True
    except Exception:
        return False


def rewrite_image_without_metadata(filepath, ext):
    """
    Yuklenen gorseli yeniden kaydederek EXIF/GPS gibi metadata bilgisini temizler.
    """
    format_map = {
        'jpg': 'JPEG',
        'jpeg': 'JPEG',
        'png': 'PNG',
        'webp': 'WEBP',
    }
    image_format = format_map.get(ext)
    if not image_format:
        return False

    try:
        with Image.open(filepath) as img:
            img.load()
            img = ImageOps.exif_transpose(img)
            if image_format == 'JPEG' and img.mode not in ('RGB', 'L'):
                img = img.convert('RGB')
            elif image_format in {'PNG', 'WEBP'} and img.mode == 'P':
                img = img.convert('RGBA')

            save_kwargs = {}
            if image_format in {'JPEG', 'WEBP'}:
                save_kwargs.update({'quality': 88, 'optimize': True})
            elif image_format == 'PNG':
                save_kwargs.update({'optimize': True})

            img.save(filepath, format=image_format, **save_kwargs)
        return True
    except Exception:
        return False


def save_upload_with_gps(file, public=False):
    """
    Yuklenen dosyayi kaydeder.
    - secure_filename ile dosya adini temizler.
    - Uzantiyi kontrol eder.
    - Pillow ile gercek goruntu dogrulamasi yapar ve metadata bilgisini temizler.
    - Benzersiz isim verir.
    Donus: (medya yolu, exif_lat, exif_lng) veya (None, None, None).
    Hata durumunda flash mesaji gonderir.
    """
    if not file or not file.filename:
        return None, None, None

    # secure_filename ile guvenli ad olustur
    safe_name = secure_filename(file.filename)
    if not safe_name or not allowed_file(safe_name):
        flash('Gecersiz dosya formati. Sadece JPG, JPEG, PNG, WEBP kabul edilir.', 'error')
        return None, None, None

    ext = safe_name.rsplit('.', 1)[1].lower()
    unique_name = f"{uuid.uuid4().hex}.{ext}"
    target_folder = PUBLIC_UPLOAD_FOLDER if public else UPLOAD_FOLDER
    relative_prefix = 'public_uploads' if public else 'uploads'
    filepath = os.path.join(target_folder, unique_name)

    try:
        file.save(filepath)
    except Exception:
        flash('Dosya kaydedilirken bir hata olustu.', 'error')
        return None, None, None

    if not verify_image(filepath):
        try:
            os.remove(filepath)
        except OSError:
            pass
        flash('Yuklenen dosya gecerli bir goruntu degil.', 'error')
        return None, None, None

    exif_lat, exif_lng = get_exif_gps(filepath)
    if not rewrite_image_without_metadata(filepath, ext):
        try:
            os.remove(filepath)
        except OSError:
            pass
        flash('Yuklenen gorsel islenirken bir hata olustu.', 'error')
        return None, None, None

    return f"{relative_prefix}/{unique_name}", exif_lat, exif_lng


def save_upload(file, public=False):
    """Yuklenen gorseli kaydeder ve sadece medya yolunu dondurur."""
    path, _, _ = save_upload_with_gps(file, public=public)
    return path


def get_exif_gps(filepath):
    """
    Fotograftan EXIF GPS bilgisini okur.
    Basariliysa (latitude, longitude) dondurur, yoksa (None, None).
    """
    try:
        with Image.open(filepath) as image:
            exif_data = image._getexif()
        if not exif_data:
            return None, None

        gps_info = {}
        for tag_id, value in exif_data.items():
            tag_name = TAGS.get(tag_id, tag_id)
            if tag_name == 'GPSInfo':
                for gps_tag_id in value:
                    gps_tag_name = GPSTAGS.get(gps_tag_id, gps_tag_id)
                    gps_info[gps_tag_name] = value[gps_tag_id]

        if not gps_info:
            return None, None

        def convert_to_degrees(value):
            d = float(value[0])
            m = float(value[1])
            s = float(value[2])
            return d + (m / 60.0) + (s / 3600.0)

        if 'GPSLatitude' in gps_info and 'GPSLongitude' in gps_info:
            lat = convert_to_degrees(gps_info['GPSLatitude'])
            lng = convert_to_degrees(gps_info['GPSLongitude'])

            if gps_info.get('GPSLatitudeRef', 'N') == 'S':
                lat = -lat
            if gps_info.get('GPSLongitudeRef', 'E') == 'W':
                lng = -lng

            return lat, lng

    except Exception:
        pass

    return None, None


def generate_qr_code(hive_id, hive_type='fixed'):
    """Kovan icin QR kod uretir ve dosya yolunu dondurur."""
    if hive_type == 'swarm':
        qr_filename = f"qr_swarm_{hive_id}.png"
        endpoint = 'public_qr_swarm'
    else:
        qr_filename = f"qr_hive_{hive_id}.png"
        endpoint = 'public_qr_fixed'
    qr_filepath = os.path.join(QR_FOLDER, qr_filename)

    # QR kodun yonlendirecegi URL
    url = build_public_url(endpoint, id=hive_id)

    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#5D4037", back_color="white")
    img.save(qr_filepath)

    return f"qrcodes/{qr_filename}"


def normalize_source_type(value):
    """Mesaj kaynagi tipini desteklenen kısa bicime indirger."""
    normalized = (value or '').strip().lower()
    aliases = {
        'fixed': 'fixed',
        'fixed_hive': 'fixed',
        'sabit': 'fixed',
        'swarm': 'swarm',
        'swarm_hive': 'swarm',
        'ogul': 'swarm',
        'oğul': 'swarm',
    }
    return aliases.get(normalized)


def parse_source_id(value):
    """Mesaj kaynagi id alanini pozitif tamsayi olarak okur."""
    try:
        source_id = int(value)
    except (TypeError, ValueError):
        return None
    return source_id if source_id > 0 else None


def get_message_source_context(source_type, source_id):
    """Public mesaj formu ve admin listesi icin kovan kaynagi dondurur."""
    normalized_type = normalize_source_type(source_type)
    normalized_id = parse_source_id(source_id)
    if not normalized_type or not normalized_id:
        return None

    if normalized_type == 'fixed':
        hive = query_db('''
            SELECT fh.id, fh.kovan_no, fh.aktif, a.arilik_adi,
                   a.latitude, a.longitude
            FROM fixed_hives fh
            JOIN apiaries a ON fh.arilik_id = a.id
            WHERE fh.id = ?
        ''', (normalized_id,), one=True)
        if not hive:
            return None
        return {
            'type': 'fixed',
            'id': hive['id'],
            'title': f"Kovan {hive['kovan_no']}",
            'subtitle': 'Sabit kovan',
            'admin_label': f"Sabit kovan {hive['kovan_no']} - {hive['arilik_adi']}",
            'public_label': f"Kovan etiketi: {hive['kovan_no']}",
            'detail_endpoint': 'fixed_hive_detail',
            'location_latitude': hive['latitude'],
            'location_longitude': hive['longitude'],
        }

    hive = query_db(
        'SELECT id, ad, aktif, latitude, longitude FROM swarm_hives WHERE id = ?',
        (normalized_id,),
        one=True
    )
    if not hive:
        return None
    return {
        'type': 'swarm',
        'id': hive['id'],
        'title': hive['ad'],
        'subtitle': 'Oğul kovanı',
        'admin_label': f"Oğul kovanı {hive['ad']}",
        'public_label': f"Kovan etiketi: {hive['ad']}",
        'detail_endpoint': 'swarm_detail',
        'location_latitude': hive['latitude'],
        'location_longitude': hive['longitude'],
    }


def public_hive_context(source_context):
    """Girişsiz QR ekraninda gösterilecek güvenli kovan baglamini hazirlar."""
    detail_url = url_for(source_context['detail_endpoint'], id=source_context['id'])
    has_location = (
        source_context.get('location_latitude') is not None
        and source_context.get('location_longitude') is not None
    )
    return {
        **source_context,
        'has_location': has_location,
        'message_url': url_for(
            'public_message',
            kaynak_turu=source_context['type'],
            kaynak_id=source_context['id'],
        ),
        'login_url': url_for('login', next=detail_url),
    }


def is_overdue(son_kontrol_tarihi, threshold_days=7):
    """
    Son kontrol tarihinin belirli gun sayisindan eski olup olmadigini kontrol eder.
    - Ogul kovanlari icin varsayilan esik 7 gundur.
    - Sabit kovanlar icin farkli esik verilebilir.
    - Tarih bossa 'hic kontrol edilmedi' kabul edilir.
    """
    if not son_kontrol_tarihi:
        return True
    try:
        last_check = datetime.strptime(son_kontrol_tarihi, '%Y-%m-%d')
        return (datetime.now() - last_check).days > threshold_days
    except (ValueError, TypeError):
        return True


# ---------------------------------------------------------------------------
# Template filtreleri
# ---------------------------------------------------------------------------

@app.template_filter('format_date')
def format_date_filter(value):
    """Tarih formatlama filtresi."""
    if not value:
        return 'Belirtilmedi'
    try:
        dt = datetime.strptime(value, '%Y-%m-%d')
        return dt.strftime('%d.%m.%Y')
    except (ValueError, TypeError):
        return value


@app.template_filter('days_ago')
def days_ago_filter(value):
    """Kac gun once oldugunu hesaplar."""
    if not value:
        return 'Hiç kontrol edilmedi'
    try:
        dt = datetime.strptime(value, '%Y-%m-%d')
        days = (datetime.now() - dt).days
        if days == 0:
            return 'Bugün'
        elif days == 1:
            return 'Dün'
        else:
            return f'{days} gün önce'
    except (ValueError, TypeError):
        return 'Bilinmiyor'


@app.context_processor
def utility_processor():
    """Template'lere yardimci fonksiyonlar ekler."""
    return {
        'is_overdue': is_overdue,
        'now': datetime.now(),
        'csrf_token': csrf_token,
        'is_authenticated': session.get('authenticated', False),
        'asset_version': ASSET_VERSION,
        'media_url': media_url,
    }


# ===================================================================
# ROTALAR
# ===================================================================

# ---------------------------------------------------------------------------
# Oturum
# ---------------------------------------------------------------------------

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Basit sifre ile giris."""
    if request.method == 'POST':
        password = request.form.get('password', '')
        if secrets.compare_digest(password, APP_PASSWORD):
            next_url = request.form.get('next') or request.args.get('next') or url_for('index')
            if not is_safe_local_next(next_url):
                next_url = url_for('index')

            session.clear()
            session['authenticated'] = True
            session['_csrf_token'] = secrets.token_urlsafe(32)
            flash('Giris yapildi.', 'success')
            return redirect(next_url)

        flash('Sifre hatali.', 'error')

    next_url = request.args.get('next') or url_for('index')
    if not is_safe_local_next(next_url):
        next_url = url_for('index')
    return render_template('login.html', next_url=next_url)


@app.route('/logout', methods=['POST'])
def logout():
    """Oturumu kapatir."""
    session.clear()
    flash('Cikis yapildi.', 'success')
    return redirect(url_for('login'))


# ---------------------------------------------------------------------------
# Public Sayfalar
# ---------------------------------------------------------------------------

@app.route('/')
def public_home():
    """Halka acik bilgilendirici ana sayfa."""
    stats = query_db('''
        SELECT
            (SELECT COUNT(*) FROM apiaries) as apiary_count,
            (SELECT COUNT(*) FROM fixed_hives WHERE aktif = 1) as active_hive_count,
            (SELECT COUNT(*) FROM swarm_hives WHERE aktif = 1) as swarm_hive_count,
            (
                (SELECT COUNT(*) FROM inspections)
                + (SELECT COUNT(*) FROM swarm_inspections)
            ) as inspection_count
    ''', one=True)
    latest_posts = query_db('''
        SELECT * FROM public_posts
        WHERE yayinla = 1
        ORDER BY yayin_tarihi DESC, id DESC
        LIMIT 3
    ''')
    latest_stories = query_db('''
        SELECT * FROM honey_stories
        WHERE yayinla = 1
        ORDER BY yayin_tarihi DESC, id DESC
        LIMIT 2
    ''')
    return render_template(
        'public_home.html',
        stats=stats,
        latest_posts=latest_posts,
        latest_stories=latest_stories,
    )


@app.route('/q/fixed/<int:id>')
def public_qr_fixed(id):
    """Sabit kovan QR girisini tek public akisa yonlendirir."""
    return redirect(url_for('fixed_hive_detail', id=id))


@app.route('/q/swarm/<int:id>')
def public_qr_swarm(id):
    """Ogul kovani QR girisini tek public akisa yonlendirir."""
    return redirect(url_for('swarm_detail', id=id))


@app.route('/mesaj-birak', methods=['GET', 'POST'])
def public_message():
    """Ziyaretcilerden mesaj alir."""
    source_context = get_message_source_context(
        request.values.get('kaynak_turu') or request.values.get('source_type'),
        request.values.get('kaynak_id') or request.values.get('source_id'),
    )
    form = {
        'ad': '',
        'iletisim': '',
        'kategori': HIVE_MESSAGE_CATEGORIES[0] if source_context else 'Diğer',
        'konu': '',
        'mesaj': '',
        'kaynak_turu': source_context['type'] if source_context else '',
        'kaynak_id': str(source_context['id']) if source_context else '',
    }

    if request.method == 'POST':
        # Basit bot filtresi; dolu gelirse sessizce basarili gibi davran.
        if request.form.get('website', '').strip():
            flash('Mesajiniz alindi.', 'success')
            return redirect(url_for('public_message'))

        defaults = form.copy()
        form = {
            key: request.form.get(key, defaults[key]).strip()
            for key in defaults
        }
        source_context = get_message_source_context(
            form['kaynak_turu'],
            form['kaynak_id'],
        )
        if source_context:
            form['kaynak_turu'] = source_context['type']
            form['kaynak_id'] = str(source_context['id'])

        errors = []
        if not form['mesaj']:
            errors.append('Mesaj alanı zorunludur.')
        if form['kategori'] not in PUBLIC_MESSAGE_CATEGORIES:
            errors.append('Konu türü geçersiz.')
        if (form['kaynak_turu'] or form['kaynak_id']) and not source_context:
            errors.append('Mesaj kaynağı geçersiz.')
        if len(form['ad']) > 80:
            errors.append('Ad en fazla 80 karakter olabilir.')
        if len(form['iletisim']) > 120:
            errors.append('İletişim bilgisi en fazla 120 karakter olabilir.')
        if len(form['konu']) > 120:
            errors.append('Konu en fazla 120 karakter olabilir.')
        if len(form['mesaj']) > 2000:
            errors.append('Mesaj en fazla 2000 karakter olabilir.')

        if errors:
            flash_errors(errors)
        else:
            execute_db('''
                INSERT INTO public_messages
                (ad, iletisim, kategori, konu, mesaj, ip_adresi, user_agent,
                 kaynak_turu, kaynak_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                form['ad'] or 'İsimsiz ziyaretçi',
                form['iletisim'] or None,
                form['kategori'],
                form['konu'] or None,
                form['mesaj'],
                request.remote_addr,
                request.headers.get('User-Agent', '')[:255],
                source_context['type'] if source_context else None,
                source_context['id'] if source_context else None,
            ))
            flash('Mesajiniz alindi.', 'success')
            if source_context:
                return redirect(url_for(
                    'public_message',
                    kaynak_turu=source_context['type'],
                    kaynak_id=source_context['id'],
                ))
            return redirect(url_for('public_message'))

    categories = HIVE_MESSAGE_CATEGORIES if source_context else GENERAL_MESSAGE_CATEGORIES
    return render_template(
        'public_message.html',
        form=form,
        categories=categories,
        source_context=source_context,
    )


@app.route('/gunluk')
def public_journal():
    """Public sezon gunlugu yazilari."""
    category = request.args.get('kategori', '').strip()
    params = []
    where = 'WHERE yayinla = 1'
    if category in PUBLIC_POST_CATEGORIES:
        where += ' AND kategori = ?'
        params.append(category)

    posts = query_db(f'''
        SELECT * FROM public_posts
        {where}
        ORDER BY yayin_tarihi DESC, id DESC
    ''', params)
    return render_template(
        'public_journal.html',
        posts=posts,
        categories=PUBLIC_POST_CATEGORIES,
        selected_category=category,
    )


@app.route('/bal-hikayeleri')
def public_honey_stories():
    """Public bal hikayeleri listesi."""
    stories = query_db('''
        SELECT * FROM honey_stories
        WHERE yayinla = 1
        ORDER BY yayin_tarihi DESC, id DESC
    ''')
    return render_template('public_honey_stories.html', stories=stories)


@app.route('/bal-hikayeleri/<int:id>')
def public_honey_story_detail(id):
    """Public bal hikayesi detayi."""
    story = query_db('''
        SELECT * FROM honey_stories
        WHERE id = ? AND yayinla = 1
    ''', (id,), one=True)
    if not story:
        abort(404)
    return render_template('public_honey_story_detail.html', story=story)


@app.route('/offline')
def offline():
    """PWA offline durumunda gosterilecek sade sayfa."""
    return render_template('offline.html')


@app.route('/manifest.webmanifest')
def manifest():
    """PWA manifest dosyasi."""
    return app.send_static_file('manifest.webmanifest')


@app.route('/sw.js')
def service_worker():
    """Service worker dosyasi."""
    response = app.send_static_file('sw.js')
    response.headers['Cache-Control'] = 'no-cache'
    return response


@app.route('/public-media/<path:filename>')
def public_upload(filename):
    """Public icerik gorsellerini servis eder."""
    return send_from_directory(PUBLIC_UPLOAD_FOLDER, filename)


@app.route('/media/uploads/<path:filename>')
def private_upload(filename):
    """Kovan ve kontrol fotograflarini oturum arkasinda servis eder."""
    if not os.path.exists(os.path.join(UPLOAD_FOLDER, filename)):
        legacy_folder = os.path.join(BASE_DIR, 'static', 'uploads')
        return send_from_directory(legacy_folder, filename)
    return send_from_directory(UPLOAD_FOLDER, filename)


@app.route('/media/qrcodes/<path:filename>')
def private_qrcode(filename):
    """QR kod dosyalarini oturum arkasinda servis eder."""
    if not os.path.exists(os.path.join(QR_FOLDER, filename)):
        legacy_folder = os.path.join(BASE_DIR, 'static', 'qrcodes')
        return send_from_directory(legacy_folder, filename)
    return send_from_directory(QR_FOLDER, filename)


# ---------------------------------------------------------------------------
# Admin Ana Sayfa (Harita)
# ---------------------------------------------------------------------------

@app.route('/admin')
def index():
    """Admin harita sayfasi."""
    return render_template('index.html')


@app.route('/admin/backups')
def admin_backups():
    """Yedek dosyalarini listeler."""
    backups = list_backup_files()
    return render_template('admin_backups.html', backups=backups)


@app.route('/admin/backups/create', methods=['POST'])
def admin_backup_create():
    """Yeni zip yedegi olusturur."""
    try:
        filename = create_backup_zip()
    except Exception as exc:
        app.logger.exception('Yedek olusturulamadi')
        flash(f'Yedek oluşturulamadı: {exc}', 'error')
    else:
        flash(f'Yeni yedek oluşturuldu: {filename}', 'success')
    return redirect(url_for('admin_backups'))


@app.route('/admin/backups/<path:filename>/download')
def admin_backup_download(filename):
    """Yedek dosyasini oturum arkasinda indirir."""
    path = backup_path_for(filename)
    if not path or not os.path.exists(path):
        abort(404)
    return send_from_directory(BACKUP_FOLDER, filename, as_attachment=True, download_name=filename)


@app.route('/admin/backups/<path:filename>/delete', methods=['POST'])
def admin_backup_delete(filename):
    """Yedek dosyasini siler."""
    path = backup_path_for(filename)
    if not path or not os.path.exists(path):
        flash('Yedek dosyası bulunamadı.', 'error')
    else:
        os.remove(path)
        flash('Yedek dosyası silindi.', 'success')
    return redirect(url_for('admin_backups'))


@app.route('/admin/export')
def admin_export():
    """Disa aktarma indirme sayfasi."""
    return render_template('admin_export.html', datasets=export_dataset_options())


@app.route('/admin/export/swarm-hives.csv')
def admin_export_swarm_hives_csv():
    headers, rows = swarm_export_data()
    filename = f'ogul_kovanlari_{export_date_suffix()}.csv'
    return csv_download_response(filename, headers, rows)


@app.route('/admin/export/swarm-hives.xlsx')
def admin_export_swarm_hives_xlsx():
    return single_sheet_excel_response('Oğul Kovanları', 'ogul_kovanlari', swarm_export_data)


@app.route('/admin/export/fixed-hives.csv')
def admin_export_fixed_hives_csv():
    headers, rows = fixed_hive_export_data()
    filename = f'sabit_kovanlar_{export_date_suffix()}.csv'
    return csv_download_response(filename, headers, rows)


@app.route('/admin/export/fixed-hives.xlsx')
def admin_export_fixed_hives_xlsx():
    return single_sheet_excel_response('Sabit Kovanlar', 'sabit_kovanlar', fixed_hive_export_data)


@app.route('/admin/export/inspections.csv')
def admin_export_inspections_csv():
    headers, rows = inspection_export_data()
    filename = f'kontrol_kayitlari_{export_date_suffix()}.csv'
    return csv_download_response(filename, headers, rows)


@app.route('/admin/export/inspections.xlsx')
def admin_export_inspections_xlsx():
    return single_sheet_excel_response('Kontrol Kayıtları', 'kontrol_kayitlari', inspection_export_data)


@app.route('/admin/export/apiary-summary.csv')
def admin_export_apiary_summary_csv():
    headers, rows = apiary_summary_export_data()
    filename = f'arilik_ozeti_{export_date_suffix()}.csv'
    return csv_download_response(filename, headers, rows)


@app.route('/admin/export/apiary-summary.xlsx')
def admin_export_apiary_summary_xlsx():
    return single_sheet_excel_response('Arılık Özeti', 'arilik_ozeti', apiary_summary_excel_data)


@app.route('/admin/export/all.xlsx')
def admin_export_all_xlsx():
    return excel_download_response()


@app.route('/admin/messages')
def admin_messages():
    """Public mesajlari listeler."""
    status = request.args.get('durum', '').strip()
    category = request.args.get('kategori', '').strip()
    search = request.args.get('q', '').strip()
    source_label_sql = """
        CASE
            WHEN pm.kaynak_turu = 'fixed' AND fh.id IS NOT NULL
                THEN 'Sabit kovan ' || fh.kovan_no || ' - ' || a.arilik_adi
            WHEN pm.kaynak_turu = 'fixed'
                THEN 'Sabit kovan #' || pm.kaynak_id
            WHEN pm.kaynak_turu = 'swarm' AND sh.id IS NOT NULL
                THEN 'Oğul kovanı ' || sh.ad
            WHEN pm.kaynak_turu = 'swarm'
                THEN 'Oğul kovanı #' || pm.kaynak_id
            ELSE NULL
        END
    """
    params = []
    conditions = []
    if status in PUBLIC_MESSAGE_STATUSES:
        conditions.append('pm.durum = ?')
        params.append(status)
    if category in PUBLIC_MESSAGE_CATEGORIES:
        conditions.append('pm.kategori = ?')
        params.append(category)
    if search:
        like = f"%{search.lower()}%"
        conditions.append(f'''
            (
                lower(pm.ad) LIKE ?
                OR lower(COALESCE(pm.iletisim, '')) LIKE ?
                OR lower(COALESCE(pm.konu, '')) LIKE ?
                OR lower(pm.mesaj) LIKE ?
                OR lower(COALESCE(pm.yanit_notu, '')) LIKE ?
                OR lower(COALESCE({source_label_sql}, '')) LIKE ?
            )
        ''')
        params.extend([like, like, like, like, like, like])
    where_sql = f"WHERE {' AND '.join(conditions)}" if conditions else ''

    messages = query_db(f'''
        SELECT pm.*, {source_label_sql} AS kaynak_label
        FROM public_messages pm
        LEFT JOIN fixed_hives fh
            ON pm.kaynak_turu = 'fixed' AND pm.kaynak_id = fh.id
        LEFT JOIN apiaries a
            ON fh.arilik_id = a.id
        LEFT JOIN swarm_hives sh
            ON pm.kaynak_turu = 'swarm' AND pm.kaynak_id = sh.id
        {where_sql}
        ORDER BY pm.okundu ASC, pm.olusturma_tarihi DESC
    ''', params)
    unread_count = query_db(
        'SELECT COUNT(*) as count FROM public_messages WHERE okundu = 0',
        one=True
    )
    return render_template(
        'public_messages.html',
        messages=messages,
        unread_count=unread_count['count'] if unread_count else 0,
        statuses=PUBLIC_MESSAGE_STATUSES,
        categories=PUBLIC_MESSAGE_CATEGORIES,
        selected_status=status,
        selected_category=category,
        search=search,
    )


@app.route('/admin/messages/<int:id>/read', methods=['POST'])
def admin_message_mark_read(id):
    """Mesaji okundu olarak isaretler."""
    execute_db(
        """UPDATE public_messages
           SET okundu = 1, durum = 'Okundu', okundu_tarihi = datetime('now', 'localtime')
           WHERE id = ?""",
        (id,)
    )
    flash('Mesaj okundu olarak işaretlendi.', 'success')
    return redirect_to_next('admin_messages')


@app.route('/admin/messages/<int:id>/status', methods=['POST'])
def admin_message_update_status(id):
    """Mesaj durumunu ve yanit notunu gunceller."""
    status = request.form.get('durum', 'Yeni')
    note = request.form.get('yanit_notu', '').strip()
    if status not in PUBLIC_MESSAGE_STATUSES:
        flash('Mesaj durumu geçersiz.', 'error')
        return redirect_to_next('admin_messages')

    read = 0 if status == 'Yeni' else 1
    read_at_sql = ", okundu_tarihi = COALESCE(okundu_tarihi, datetime('now', 'localtime'))" if read else ", okundu_tarihi = NULL"
    execute_db(f'''
        UPDATE public_messages
        SET durum = ?, yanit_notu = ?, okundu = ?{read_at_sql}
        WHERE id = ?
    ''', (status, note or None, read, id))
    flash('Mesaj durumu güncellendi.', 'success')
    return redirect_to_next('admin_messages')


@app.route('/admin/content')
def admin_content():
    """Public icerik yonetimi ana sayfasi."""
    post_count = query_db('SELECT COUNT(*) as count FROM public_posts', one=True)
    story_count = query_db('SELECT COUNT(*) as count FROM honey_stories', one=True)
    return render_template(
        'admin_content.html',
        post_count=post_count['count'] if post_count else 0,
        story_count=story_count['count'] if story_count else 0,
    )


@app.route('/admin/posts')
def admin_posts():
    """Sezon gunlugu yazilarini listeler."""
    posts = query_db('SELECT * FROM public_posts ORDER BY yayin_tarihi DESC, id DESC')
    return render_template('admin_posts.html', posts=posts)


@app.route('/admin/posts/new', methods=['GET', 'POST'])
def admin_post_new():
    """Yeni sezon gunlugu yazisi."""
    if request.method == 'POST':
        baslik = request.form.get('baslik', '').strip()
        kategori = request.form.get('kategori', 'Duyuru')
        ozet = request.form.get('ozet', '').strip()
        icerik = request.form.get('icerik', '').strip()
        yayin_tarihi = request.form.get('yayin_tarihi', '').strip()
        yayinla = 1 if request.form.get('yayinla') else 0

        yayin_tarihi, date_error = validate_date(yayin_tarihi, 'Yayın tarihi')
        errors = [date_error]
        if not baslik:
            errors.append('Başlık zorunludur.')
        if not icerik:
            errors.append('İçerik zorunludur.')
        if kategori not in PUBLIC_POST_CATEGORIES:
            errors.append('Kategori geçersiz.')
        if any(errors):
            flash_errors(errors)
            return render_template('admin_post_form.html', post=None, categories=PUBLIC_POST_CATEGORIES)

        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu = save_upload(file, public=True)
            if not fotograf_yolu:
                post_form = {
                    'baslik': baslik,
                    'kategori': kategori,
                    'ozet': ozet,
                    'icerik': icerik,
                    'yayin_tarihi': yayin_tarihi or datetime.now().strftime('%Y-%m-%d'),
                    'yayinla': yayinla,
                }
                return render_template('admin_post_form.html', post=post_form,
                                       categories=PUBLIC_POST_CATEGORIES)

        execute_db('''
            INSERT INTO public_posts
            (baslik, kategori, ozet, icerik, fotograf_yolu, yayinla, yayin_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (baslik, kategori, ozet or None, icerik, fotograf_yolu, yayinla,
              yayin_tarihi or datetime.now().strftime('%Y-%m-%d')))
        flash('Günlük yazısı eklendi.', 'success')
        return redirect(url_for('admin_posts'))

    return render_template('admin_post_form.html', post=None, categories=PUBLIC_POST_CATEGORIES)


@app.route('/admin/posts/<int:id>/edit', methods=['GET', 'POST'])
def admin_post_edit(id):
    """Sezon gunlugu yazisini duzenler."""
    post = query_db('SELECT * FROM public_posts WHERE id = ?', (id,), one=True)
    if not post:
        flash('Yazı bulunamadı.', 'error')
        return redirect(url_for('admin_posts'))

    if request.method == 'POST':
        baslik = request.form.get('baslik', '').strip()
        kategori = request.form.get('kategori', 'Duyuru')
        ozet = request.form.get('ozet', '').strip()
        icerik = request.form.get('icerik', '').strip()
        yayin_tarihi = request.form.get('yayin_tarihi', '').strip()
        yayinla = 1 if request.form.get('yayinla') else 0

        yayin_tarihi, date_error = validate_date(yayin_tarihi, 'Yayın tarihi')
        errors = [date_error]
        if not baslik:
            errors.append('Başlık zorunludur.')
        if not icerik:
            errors.append('İçerik zorunludur.')
        if kategori not in PUBLIC_POST_CATEGORIES:
            errors.append('Kategori geçersiz.')
        if any(errors):
            flash_errors(errors)
            return render_template('admin_post_form.html', post=post, categories=PUBLIC_POST_CATEGORIES)

        fotograf_yolu = post['fotograf_yolu']
        file = request.files.get('fotograf')
        if file and file.filename:
            new_photo = save_upload(file, public=True)
            if not new_photo:
                return render_template('admin_post_form.html', post=post,
                                       categories=PUBLIC_POST_CATEGORIES)
            fotograf_yolu = new_photo

        execute_db('''
            UPDATE public_posts
            SET baslik=?, kategori=?, ozet=?, icerik=?, fotograf_yolu=?, yayinla=?,
                yayin_tarihi=?, guncelleme_tarihi=datetime('now', 'localtime')
            WHERE id=?
        ''', (baslik, kategori, ozet or None, icerik, fotograf_yolu, yayinla,
              yayin_tarihi or datetime.now().strftime('%Y-%m-%d'), id))
        flash('Günlük yazısı güncellendi.', 'success')
        return redirect(url_for('admin_posts'))

    return render_template('admin_post_form.html', post=post, categories=PUBLIC_POST_CATEGORIES)


@app.route('/admin/honey-stories')
def admin_honey_stories():
    """Bal hikayelerini listeler."""
    stories = query_db('SELECT * FROM honey_stories ORDER BY yayin_tarihi DESC, id DESC')
    return render_template('admin_honey_stories.html', stories=stories)


@app.route('/admin/honey-stories/new', methods=['GET', 'POST'])
def admin_honey_story_new():
    """Yeni bal hikayesi."""
    if request.method == 'POST':
        data = {
            'baslik': request.form.get('baslik', '').strip(),
            'hasat_donemi': request.form.get('hasat_donemi', '').strip(),
            'bolge_notu': request.form.get('bolge_notu', '').strip(),
            'flora_notu': request.form.get('flora_notu', '').strip(),
            'tadim_notu': request.form.get('tadim_notu', '').strip(),
            'saklama_notu': request.form.get('saklama_notu', '').strip(),
            'yayin_tarihi': request.form.get('yayin_tarihi', '').strip(),
        }
        data['yayin_tarihi'], date_error = validate_date(data['yayin_tarihi'], 'Yayın tarihi')
        errors = [date_error]
        if not data['baslik']:
            errors.append('Başlık zorunludur.')
        if any(errors):
            flash_errors(errors)
            return render_template('admin_honey_story_form.html', story=data)

        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu = save_upload(file, public=True)
            if not fotograf_yolu:
                data['yayinla'] = 1 if request.form.get('yayinla') else 0
                return render_template('admin_honey_story_form.html', story=data)

        execute_db('''
            INSERT INTO honey_stories
            (baslik, hasat_donemi, bolge_notu, flora_notu, tadim_notu, saklama_notu,
             fotograf_yolu, yayinla, yayin_tarihi)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['baslik'], data['hasat_donemi'] or None, data['bolge_notu'] or None,
            data['flora_notu'] or None, data['tadim_notu'] or None,
            data['saklama_notu'] or None, fotograf_yolu,
            1 if request.form.get('yayinla') else 0,
            data['yayin_tarihi'] or datetime.now().strftime('%Y-%m-%d'),
        ))
        flash('Bal hikayesi eklendi.', 'success')
        return redirect(url_for('admin_honey_stories'))

    return render_template('admin_honey_story_form.html', story=None)


@app.route('/admin/honey-stories/<int:id>/edit', methods=['GET', 'POST'])
def admin_honey_story_edit(id):
    """Bal hikayesini duzenler."""
    story = query_db('SELECT * FROM honey_stories WHERE id = ?', (id,), one=True)
    if not story:
        flash('Bal hikayesi bulunamadı.', 'error')
        return redirect(url_for('admin_honey_stories'))

    if request.method == 'POST':
        data = {
            'baslik': request.form.get('baslik', '').strip(),
            'hasat_donemi': request.form.get('hasat_donemi', '').strip(),
            'bolge_notu': request.form.get('bolge_notu', '').strip(),
            'flora_notu': request.form.get('flora_notu', '').strip(),
            'tadim_notu': request.form.get('tadim_notu', '').strip(),
            'saklama_notu': request.form.get('saklama_notu', '').strip(),
            'yayin_tarihi': request.form.get('yayin_tarihi', '').strip(),
        }
        data['yayin_tarihi'], date_error = validate_date(data['yayin_tarihi'], 'Yayın tarihi')
        errors = [date_error]
        if not data['baslik']:
            errors.append('Başlık zorunludur.')
        if any(errors):
            flash_errors(errors)
            return render_template('admin_honey_story_form.html', story=story)

        fotograf_yolu = story['fotograf_yolu']
        file = request.files.get('fotograf')
        if file and file.filename:
            new_photo = save_upload(file, public=True)
            if not new_photo:
                return render_template('admin_honey_story_form.html', story=story)
            fotograf_yolu = new_photo

        execute_db('''
            UPDATE honey_stories
            SET baslik=?, hasat_donemi=?, bolge_notu=?, flora_notu=?, tadim_notu=?,
                saklama_notu=?, fotograf_yolu=?, yayinla=?, yayin_tarihi=?,
                guncelleme_tarihi=datetime('now', 'localtime')
            WHERE id=?
        ''', (
            data['baslik'], data['hasat_donemi'] or None, data['bolge_notu'] or None,
            data['flora_notu'] or None, data['tadim_notu'] or None,
            data['saklama_notu'] or None, fotograf_yolu,
            1 if request.form.get('yayinla') else 0,
            data['yayin_tarihi'] or datetime.now().strftime('%Y-%m-%d'), id,
        ))
        flash('Bal hikayesi güncellendi.', 'success')
        return redirect(url_for('admin_honey_stories'))

    return render_template('admin_honey_story_form.html', story=story)


@app.route('/admin/hives')
def admin_hives():
    """Ogul ve sabit kovanlari tablo halinde listeler."""
    hive_type = request.args.get('tip', 'all')
    active_filter = request.args.get('aktif', 'all')
    query = request.args.get('q', '').strip()
    if hive_type not in {'all', 'fixed', 'swarm'}:
        hive_type = 'all'
    if active_filter not in {'all', '0', '1'}:
        active_filter = 'all'

    fixed_hives = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        ORDER BY a.arilik_adi, fh.sira_no, fh.konum_no
    ''')
    swarm_hives = query_db('''
        SELECT * FROM swarm_hives
        ORDER BY aktif DESC, son_kontrol_tarihi DESC, ad
    ''')

    return render_template(
        'admin_hives.html',
        fixed_hives=fixed_hives,
        swarm_hives=swarm_hives,
        total_count=len(fixed_hives) + len(swarm_hives),
        selected_type=hive_type,
        active_filter=active_filter,
        query=query,
        fixed_statuses=FIXED_HIVE_STATUSES,
        swarm_statuses=SWARM_STATUSES,
    )


@app.route('/admin/hives/swarm/<int:id>/quick-update', methods=['POST'])
def admin_swarm_quick_update(id):
    """Tablodan ogul kovani hizli guncelleme."""
    durum = request.form.get('durum', 'Boş')
    son_kontrol_tarihi = request.form.get('son_kontrol_tarihi', '').strip()
    aktif = 1 if request.form.get('aktif') else 0
    durum, status_error = validate_choice(durum, SWARM_STATUSES, 'Durum')
    son_kontrol_tarihi, date_error = validate_date(son_kontrol_tarihi, 'Son kontrol tarihi')
    if status_error or date_error:
        flash_errors([status_error, date_error])
    else:
        execute_db('''
            UPDATE swarm_hives
            SET durum=?, son_kontrol_tarihi=?, aktif=?, guncelleme_tarihi=datetime('now', 'localtime')
            WHERE id=?
        ''', (durum, son_kontrol_tarihi or None, aktif, id))
        if son_kontrol_tarihi:
            create_swarm_inspection_record(
                id,
                son_kontrol_tarihi,
                durum,
                son_kontrol_tarihi if durum in SWARM_ARRIVAL_STATUSES else None,
                'Hızlı güncelleme ekranından kaydedildi.'
            )
        flash('Oğul kovanı güncellendi.', 'success')
    return redirect_to_next('admin_hives')


@app.route('/admin/hives/fixed/<int:id>/quick-update', methods=['POST'])
def admin_fixed_quick_update(id):
    """Tablodan sabit kovan hizli guncelleme."""
    durum = request.form.get('durum', 'Orta')
    son_kontrol_tarihi = request.form.get('son_kontrol_tarihi', '').strip()
    aktif = 1 if request.form.get('aktif') else 0
    durum, status_error = validate_choice(durum, FIXED_HIVE_STATUSES, 'Durum')
    son_kontrol_tarihi, date_error = validate_date(son_kontrol_tarihi, 'Son kontrol tarihi')
    if status_error or date_error:
        flash_errors([status_error, date_error])
    else:
        execute_db('''
            UPDATE fixed_hives
            SET durum=?, son_kontrol_tarihi=?, aktif=?, guncelleme_tarihi=datetime('now', 'localtime')
            WHERE id=?
        ''', (durum, son_kontrol_tarihi or None, aktif, id))
        flash('Sabit kovan güncellendi.', 'success')
    return redirect_to_next('admin_hives')


@app.route('/swarm-hives/<int:id>/delete', methods=['POST'])
def swarm_delete(id):
    """Ogul kovani kaydini siler."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        flash('Oğul kovanı bulunamadı.', 'error')
        return redirect(url_for('swarm_list'))
    if not delete_confirmation_valid():
        flash('Silme onayı geçersiz. Silmek için uyarıda SIL yazmanız gerekir.', 'error')
        return redirect(url_for('swarm_detail', id=id))

    execute_db('DELETE FROM swarm_hives WHERE id = ?', (id,))
    flash(f"{hive['ad']} silindi. Bağlı oğul kontrol kayıtları da silindi.", 'success')
    return redirect(url_for('swarm_list'))


@app.route('/fixed-hives/<int:id>/delete', methods=['POST'])
def fixed_hive_delete(id):
    """Sabit kovani ve bagli kontrol kayitlarini siler."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        flash('Sabit kovan bulunamadı.', 'error')
        return redirect(url_for('apiary_list'))
    if not delete_confirmation_valid():
        flash('Silme onayı geçersiz. Silmek için uyarıda SIL yazmanız gerekir.', 'error')
        return redirect(url_for('fixed_hive_detail', id=id))

    apiary_id = hive['arilik_id']
    execute_db('DELETE FROM fixed_hives WHERE id = ?', (id,))
    flash(f"{hive['kovan_no']} silindi. Bağlı kontrol kayıtları da silindi.", 'success')
    return redirect(url_for('apiary_detail', id=apiary_id))


@app.route('/apiaries/<int:id>/delete', methods=['POST'])
def apiary_delete(id):
    """Ariligi ve bagli sabit kovan/kontrol kayitlarini siler."""
    apiary = query_db('SELECT * FROM apiaries WHERE id = ?', (id,), one=True)
    if not apiary:
        flash('Arılık bulunamadı.', 'error')
        return redirect(url_for('apiary_list'))
    if not delete_confirmation_valid():
        flash('Silme onayı geçersiz. Silmek için uyarıda SIL yazmanız gerekir.', 'error')
        return redirect(url_for('apiary_detail', id=id))

    stats = query_db('''
        SELECT
            COUNT(DISTINCT fh.id) as hive_count,
            COUNT(i.id) as inspection_count
        FROM fixed_hives fh
        LEFT JOIN inspections i ON i.kovan_id = fh.id
        WHERE fh.arilik_id = ?
    ''', (id,), one=True)
    execute_db('DELETE FROM apiaries WHERE id = ?', (id,))
    hive_count = stats['hive_count'] if stats else 0
    inspection_count = stats['inspection_count'] if stats else 0
    flash(
        f"{apiary['arilik_adi']} silindi. {hive_count} sabit kovan ve "
        f"{inspection_count} kontrol kaydı da silindi.",
        'success'
    )
    return redirect(url_for('apiary_list'))


# ---------------------------------------------------------------------------
# API: Harita Verileri
# ---------------------------------------------------------------------------

@app.route('/api/map-data')
def api_map_data():
    """Haritadaki tum kovan ve arilik verilerini JSON olarak dondurur."""
    focus_type, focus_id = parse_map_focus(request.args.get('focus', ''))
    data = {
        'swarm_hives': [],
        'apiaries': [],
        'fixed_hives': []
    }

    # Ogul kovanlari
    swarms = query_db('SELECT * FROM swarm_hives')
    for s in swarms:
        if not s['latitude'] or not s['longitude']:
            continue
        foto = None
        if s['fotograf_yolu']:
            foto = media_url(s['fotograf_yolu'])
        data['swarm_hives'].append({
            'id': s['id'],
            'ad': s['ad'],
            'latitude': s['latitude'],
            'longitude': s['longitude'],
            'durum': s['durum'],
            'son_kontrol_tarihi': s['son_kontrol_tarihi'],
            'aktif': s['aktif'],
            'fotograf': foto,
            'overdue': is_overdue(s['son_kontrol_tarihi'], threshold_days=7),
            'detail_url': url_for('swarm_detail', id=s['id']),
            'focused': focus_type == 'swarm' and focus_id == s['id'],
        })

    # Ariliklar - toplam/kontrol gereken/pasif sayilari ile
    apiaries_raw = query_db('SELECT * FROM apiaries')
    for a in apiaries_raw:
        if not a['latitude'] or not a['longitude']:
            continue

        # Bu ariliga ait kovan istatistikleri
        stats = query_db('''
            SELECT
                COUNT(*) as toplam,
                SUM(CASE WHEN aktif = 0 OR durum = 'Pasif' THEN 1 ELSE 0 END) as pasif,
                SUM(CASE WHEN aktif = 1 AND durum != 'Pasif' AND (
                    durum IN ('Kontrol gerekli', 'Hastalık şüphesi var')
                    OR son_kontrol_tarihi IS NULL
                    OR julianday('now', 'localtime') - julianday(son_kontrol_tarihi) > 7
                ) THEN 1 ELSE 0 END) as kontrol_gereken
            FROM fixed_hives WHERE arilik_id = ?
        ''', (a['id'],), one=True)

        data['apiaries'].append({
            'id': a['id'],
            'arilik_adi': a['arilik_adi'],
            'latitude': a['latitude'],
            'longitude': a['longitude'],
            'aciklama': a['aciklama'],
            'toplam_kovan_sayisi': stats['toplam'] if stats else 0,
            'kontrol_gereken_sayisi': stats['kontrol_gereken'] if stats else 0,
            'pasif_sayisi': stats['pasif'] if stats else 0,
            'detail_url': url_for('apiary_detail', id=a['id']),
            'focused': focus_type == 'apiary' and focus_id == a['id'],
        })

    # Sabit kovanlar - sadece kontrol gereken veya pasif olanlari dondur
    # (haritada bunlar circleMarker ile arilik etrafinda offset ile gosterilecek)
    fixed_focus_sql = ''
    fixed_params = []
    if focus_type == 'fixed' and focus_id is not None:
        fixed_focus_sql = ' OR fh.id = ?'
        fixed_params.append(focus_id)
    fixed = query_db('''
        SELECT fh.*, a.latitude as a_lat, a.longitude as a_lng, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE a.latitude IS NOT NULL AND a.longitude IS NOT NULL
        AND ((
            fh.aktif = 0
            OR fh.durum = 'Pasif'
            OR fh.durum = 'Kontrol gerekli'
            OR fh.durum IN ('Hastalık şüphesi var', 'Zayıf', 'Ana arı sorunu var')
            OR fh.son_kontrol_tarihi IS NULL
            OR julianday('now', 'localtime') - julianday(fh.son_kontrol_tarihi) > 7
        )''' + fixed_focus_sql + ''')
    ''', tuple(fixed_params))
    for f in fixed:
        data['fixed_hives'].append({
            'id': f['id'],
            'kovan_no': f['kovan_no'],
            'arilik_id': f['arilik_id'],
            'arilik_adi': f['arilik_adi'],
            'latitude': f['a_lat'],
            'longitude': f['a_lng'],
            'durum': f['durum'],
            'son_kontrol_tarihi': f['son_kontrol_tarihi'],
            'aktif': f['aktif'],
            'overdue': is_overdue(f['son_kontrol_tarihi'], threshold_days=7),
            'detail_url': url_for('fixed_hive_detail', id=f['id']),
            'focused': focus_type == 'fixed' and focus_id == f['id'],
        })

    return jsonify(data)


# ---------------------------------------------------------------------------
# Modul 1: Ogul Kovanlari
# ---------------------------------------------------------------------------

@app.route('/swarm-hives')
def swarm_list():
    """Ogul kovanlarini listeler."""
    hives = query_db('SELECT * FROM swarm_hives ORDER BY aktif DESC, son_kontrol_tarihi DESC')
    return render_template('swarm_list.html', hives=hives)


@app.route('/swarm-hives/new', methods=['GET', 'POST'])
def swarm_new():
    """Yeni ogul kovani ekleme."""
    if request.method == 'POST':
        ad = request.form.get('ad', '').strip()
        latitude_str = request.form.get('latitude', '').strip()
        longitude_str = request.form.get('longitude', '').strip()
        kurulum_tarihi = request.form.get('kurulum_tarihi', '').strip()
        durum = request.form.get('durum', 'Boş')
        ulasim_notu = request.form.get('ulasim_notu', '').strip()
        genel_not = request.form.get('genel_not', '').strip()

        if not ad:
            flash('Kovan adı zorunludur.', 'error')
            return render_template('swarm_form.html', hive=None, statuses=SWARM_STATUSES, edit=False)

        durum, status_error = validate_choice(durum, SWARM_STATUSES, 'Durum')
        kurulum_tarihi, date_error = validate_date(kurulum_tarihi, 'Kurulum tarihi')
        if status_error or date_error:
            flash_errors([status_error, date_error])
            return render_template('swarm_form.html', hive=None, statuses=SWARM_STATUSES, edit=False)

        # Fotograf yukleme
        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu, exif_lat, exif_lng = save_upload_with_gps(file)
            if not fotograf_yolu:
                return render_template('swarm_form.html', hive=None, statuses=SWARM_STATUSES, edit=False)
            # EXIF GPS kontrol (sadece koordinat girilmediyse)
            if not latitude_str or not longitude_str:
                if exif_lat is not None and exif_lng is not None:
                    latitude_str = str(exif_lat)
                    longitude_str = str(exif_lng)
                    flash('GPS bilgisi fotograftan otomatik alindi.', 'success')
                else:
                    flash('Fotografta GPS bilgisi bulunamadi. Koordinatlari elle girebilirsiniz.', 'info')

        # Koordinat dogrulama
        lat_val, lng_val, coord_errors = parse_coordinates(latitude_str, longitude_str)
        if coord_errors:
            for err in coord_errors:
                flash(err, 'error')
            return render_template('swarm_form.html', hive=None, statuses=SWARM_STATUSES, edit=False)

        # son_kontrol_tarihi mantigi:
        # Kullanici vermemisse kurulum tarihini kullan, o da yoksa bugunu kullan.
        son_kontrol = kurulum_tarihi or datetime.now().strftime('%Y-%m-%d')

        hive_id = execute_db('''
            INSERT INTO swarm_hives (ad, latitude, longitude, kurulum_tarihi,
            son_kontrol_tarihi, durum, ulasim_notu, genel_not, fotograf_yolu)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (ad, lat_val, lng_val, kurulum_tarihi or None,
              son_kontrol, durum, ulasim_notu or None,
              genel_not or None, fotograf_yolu))
        create_swarm_inspection_record(
            hive_id,
            son_kontrol,
            durum,
            son_kontrol if durum in SWARM_ARRIVAL_STATUSES else None,
            'İlk kayıt oluşturuldu.'
        )

        flash('Ogul kovani basariyla eklendi.', 'success')
        return redirect(url_for('swarm_list'))

    return render_template('swarm_form.html', hive=None, statuses=SWARM_STATUSES, edit=False)


@app.route('/swarm-hives/<int:id>')
def swarm_detail(id):
    """Ogul kovani detay sayfasi."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        if not session.get('authenticated'):
            flash('Kovan kaydı bulunamadı.', 'error')
            return redirect(url_for('public_home'))
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('swarm_list'))
    if not session.get('authenticated'):
        source_context = get_message_source_context('swarm', id)
        return render_template(
            'public_hive.html',
            hive_context=public_hive_context(source_context),
        )
    inspections = query_db('''
        SELECT * FROM swarm_inspections
        WHERE swarm_hive_id = ?
        ORDER BY kontrol_tarihi DESC, id DESC
        LIMIT 5
    ''', (id,))
    return render_template('swarm_detail.html', hive=hive, inspections=inspections)


@app.route('/swarm-hives/<int:id>/qr')
def swarm_hive_qr(id):
    """Ogul kovani QR kod sayfasi."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('swarm_list'))

    qr_path = generate_qr_code(id, hive_type='swarm')
    if hive['qr_kod_yolu'] != qr_path:
        execute_db('UPDATE swarm_hives SET qr_kod_yolu=? WHERE id=?', (qr_path, id))

    return render_template(
        'qr_view.html',
        hive=hive,
        qr_path=qr_path,
        qr_label=hive['ad'],
        qr_sublabel='Oğul Kovanı',
        detail_url=url_for('swarm_detail', id=id),
        public_qr_url=build_public_url('public_qr_swarm', id=id),
        back_label=f"{hive['ad']} - Oğul Kovanı",
        download_name=f"qr_ogul_{id}.png",
    )


@app.route('/swarm-hives/<int:id>/inspections/new', methods=['GET', 'POST'])
def swarm_inspection_new(id):
    """Ogul kovani icin yeni kontrol kaydi ekleme."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('swarm_list'))

    if request.method == 'POST':
        kontrol_tarihi = request.form.get('kontrol_tarihi', '').strip()
        if not kontrol_tarihi:
            kontrol_tarihi = datetime.now().strftime('%Y-%m-%d')
        durum = request.form.get('durum', hive['durum'] or 'Boş')
        ari_gelis_tarihi = request.form.get('ari_gelis_tarihi', '').strip()
        notlar = request.form.get('notlar', '').strip()

        kontrol_tarihi, kontrol_error = validate_date(kontrol_tarihi, 'Kontrol tarihi', required=True)
        durum, status_error = validate_choice(durum, SWARM_STATUSES, 'Durum')
        ari_gelis_tarihi, gelis_error = validate_date(ari_gelis_tarihi, 'Arı geliş tarihi')
        if status_error or kontrol_error or gelis_error:
            flash_errors([status_error, kontrol_error, gelis_error])
            return render_template('swarm_inspection_form.html', hive=hive, statuses=SWARM_STATUSES)

        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu = save_upload(file)
            if not fotograf_yolu:
                return render_template('swarm_inspection_form.html', hive=hive, statuses=SWARM_STATUSES)

        create_swarm_inspection_record(
            id,
            kontrol_tarihi,
            durum,
            ari_gelis_tarihi or None,
            notlar,
            fotograf_yolu
        )
        execute_db('''
            UPDATE swarm_hives
            SET durum=?, son_kontrol_tarihi=?, guncelleme_tarihi=datetime('now','localtime')
            WHERE id=?
        ''', (durum, kontrol_tarihi, id))

        flash('Oğul kontrol kaydı eklendi.', 'success')
        return redirect(url_for('swarm_detail', id=id))

    return render_template('swarm_inspection_form.html', hive=hive, statuses=SWARM_STATUSES)


@app.route('/swarm-hives/<int:id>/inspections')
def swarm_inspection_history(id):
    """Ogul kovani kontrol gecmisi."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('swarm_list'))

    inspections = query_db('''
        SELECT * FROM swarm_inspections
        WHERE swarm_hive_id = ?
        ORDER BY kontrol_tarihi DESC, id DESC
    ''', (id,))

    return render_template('swarm_inspection_history.html', hive=hive, inspections=inspections)


@app.route('/swarm-hives/<int:id>/edit', methods=['GET', 'POST'])
def swarm_edit(id):
    """Ogul kovani duzenleme."""
    hive = query_db('SELECT * FROM swarm_hives WHERE id = ?', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('swarm_list'))

    if request.method == 'POST':
        ad = request.form.get('ad', '').strip()
        latitude_str = request.form.get('latitude', '').strip()
        longitude_str = request.form.get('longitude', '').strip()
        kurulum_tarihi = request.form.get('kurulum_tarihi', '').strip()
        son_kontrol_tarihi = request.form.get('son_kontrol_tarihi', '').strip()
        durum = request.form.get('durum', 'Boş')
        ulasim_notu = request.form.get('ulasim_notu', '').strip()
        genel_not = request.form.get('genel_not', '').strip()
        aktif = 1 if request.form.get('aktif') else 0

        if not ad:
            flash('Kovan adi zorunludur.', 'error')
            return render_template('swarm_form.html', hive=hive, statuses=SWARM_STATUSES, edit=True)

        durum, status_error = validate_choice(durum, SWARM_STATUSES, 'Durum')
        kurulum_tarihi, kurulum_error = validate_date(kurulum_tarihi, 'Kurulum tarihi')
        son_kontrol_tarihi, kontrol_error = validate_date(son_kontrol_tarihi, 'Son kontrol tarihi')
        if status_error or kurulum_error or kontrol_error:
            flash_errors([status_error, kurulum_error, kontrol_error])
            return render_template('swarm_form.html', hive=hive, statuses=SWARM_STATUSES, edit=True)

        # Fotograf guncelleme
        fotograf_yolu = hive['fotograf_yolu']
        file = request.files.get('fotograf')
        if file and file.filename:
            new_foto, exif_lat, exif_lng = save_upload_with_gps(file)
            if not new_foto:
                return render_template('swarm_form.html', hive=hive, statuses=SWARM_STATUSES, edit=True)
            fotograf_yolu = new_foto
            if not latitude_str or not longitude_str:
                if exif_lat is not None and exif_lng is not None:
                    latitude_str = str(exif_lat)
                    longitude_str = str(exif_lng)
                    flash('GPS bilgisi fotograftan otomatik alindi.', 'success')
                else:
                    flash('Fotografta GPS bilgisi bulunamadi.', 'info')

        # Koordinat dogrulama
        lat_val, lng_val, coord_errors = parse_coordinates(latitude_str, longitude_str)
        if coord_errors:
            for err in coord_errors:
                flash(err, 'error')
            return render_template('swarm_form.html', hive=hive, statuses=SWARM_STATUSES, edit=True)

        execute_db('''
            UPDATE swarm_hives SET ad=?, latitude=?, longitude=?, kurulum_tarihi=?,
            son_kontrol_tarihi=?, durum=?, ulasim_notu=?, genel_not=?, fotograf_yolu=?,
            aktif=?, guncelleme_tarihi=datetime('now','localtime')
            WHERE id=?
        ''', (ad, lat_val, lng_val, kurulum_tarihi or None,
              son_kontrol_tarihi or None, durum, ulasim_notu or None,
              genel_not or None, fotograf_yolu, aktif, id))

        flash('Ogul kovani guncellendi.', 'success')
        return redirect(url_for('swarm_detail', id=id))

    return render_template('swarm_form.html', hive=hive, statuses=SWARM_STATUSES, edit=True)


# ---------------------------------------------------------------------------
# Modul 2: Sabit Ariliklar
# ---------------------------------------------------------------------------

@app.route('/apiaries')
def apiary_list():
    """Ariliklari listeler."""
    apiaries = query_db('''
        SELECT a.*, COUNT(fh.id) as kovan_sayisi
        FROM apiaries a
        LEFT JOIN fixed_hives fh ON a.id = fh.arilik_id
        GROUP BY a.id
        ORDER BY a.arilik_adi
    ''')
    return render_template('apiary_list.html', apiaries=apiaries)


@app.route('/apiaries/new', methods=['GET', 'POST'])
def apiary_new():
    """Yeni arilik ekleme."""
    if request.method == 'POST':
        arilik_adi = request.form.get('arilik_adi', '').strip()
        latitude_str = request.form.get('latitude', '').strip()
        longitude_str = request.form.get('longitude', '').strip()
        aciklama = request.form.get('aciklama', '').strip()

        if not arilik_adi:
            flash('Arilik adi zorunludur.', 'error')
            return render_template('apiary_form.html', apiary=None, edit=False)

        lat_val, lng_val, coord_errors = parse_coordinates(latitude_str, longitude_str)
        if coord_errors:
            for err in coord_errors:
                flash(err, 'error')
            return render_template('apiary_form.html', apiary=None, edit=False)

        execute_db('''
            INSERT INTO apiaries (arilik_adi, latitude, longitude, aciklama)
            VALUES (?, ?, ?, ?)
        ''', (arilik_adi, lat_val, lng_val, aciklama or None))

        flash('Arilik basariyla eklendi.', 'success')
        return redirect(url_for('apiary_list'))

    return render_template('apiary_form.html', apiary=None, edit=False)


@app.route('/apiaries/<int:id>')
def apiary_detail(id):
    """Arilik detay sayfasi ve kroki gorunumu."""
    apiary = query_db('SELECT * FROM apiaries WHERE id = ?', (id,), one=True)
    if not apiary:
        flash('Arilik bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    hives = query_db('''
        SELECT * FROM fixed_hives WHERE arilik_id = ?
        ORDER BY sira_no, konum_no
    ''', (id,))

    # Kroki duzeni icin max sira ve konum bul
    max_sira = 0
    max_konum = 0
    hive_grid = {}
    for h in hives:
        sira = h['sira_no'] or 1
        konum = h['konum_no'] or 1
        max_sira = max(max_sira, sira)
        max_konum = max(max_konum, konum)
        hive_grid[(sira, konum)] = h

    return render_template('apiary_detail.html', apiary=apiary, hives=hives,
                           hive_grid=hive_grid, max_sira=max_sira, max_konum=max_konum)


@app.route('/apiaries/<int:id>/edit', methods=['GET', 'POST'])
def apiary_edit(id):
    """Arilik duzenleme."""
    apiary = query_db('SELECT * FROM apiaries WHERE id = ?', (id,), one=True)
    if not apiary:
        flash('Arilik bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    if request.method == 'POST':
        arilik_adi = request.form.get('arilik_adi', '').strip()
        latitude_str = request.form.get('latitude', '').strip()
        longitude_str = request.form.get('longitude', '').strip()
        aciklama = request.form.get('aciklama', '').strip()

        if not arilik_adi:
            flash('Arilik adi zorunludur.', 'error')
            return render_template('apiary_form.html', apiary=apiary, edit=True)

        lat_val, lng_val, coord_errors = parse_coordinates(latitude_str, longitude_str)
        if coord_errors:
            for err in coord_errors:
                flash(err, 'error')
            return render_template('apiary_form.html', apiary=apiary, edit=True)

        execute_db('''
            UPDATE apiaries SET arilik_adi=?, latitude=?, longitude=?, aciklama=?,
            guncelleme_tarihi=datetime('now','localtime')
            WHERE id=?
        ''', (arilik_adi, lat_val, lng_val, aciklama or None, id))

        flash('Arilik guncellendi.', 'success')
        return redirect(url_for('apiary_detail', id=id))

    return render_template('apiary_form.html', apiary=apiary, edit=True)


# ---------------------------------------------------------------------------
# Modul 2: Sabit Kovanlar
# ---------------------------------------------------------------------------

@app.route('/apiaries/<int:apiary_id>/fixed-hives/new', methods=['GET', 'POST'])
def fixed_hive_new(apiary_id):
    """Ariliga yeni sabit kovan ekleme."""
    apiary = query_db('SELECT * FROM apiaries WHERE id = ?', (apiary_id,), one=True)
    if not apiary:
        flash('Arilik bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    if request.method == 'POST':
        kovan_no = request.form.get('kovan_no', '').strip()
        sira_no = request.form.get('sira_no', '1').strip()
        konum_no = request.form.get('konum_no', '1').strip()
        ana_ari_yili = request.form.get('ana_ari_yili', '').strip()
        kat_sayisi = request.form.get('kat_sayisi', '1').strip()
        cerceve_sayisi = request.form.get('cerceve_sayisi', '10').strip()
        durum = request.form.get('durum', 'Orta')
        genel_not = request.form.get('genel_not', '').strip()

        if not kovan_no:
            flash('Kovan numarasi zorunludur.', 'error')
            return render_template('fixed_hive_form.html', hive=None, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=False)

        durum, status_error = validate_choice(durum, FIXED_HIVE_STATUSES, 'Durum')
        sira_val, sira_error = parse_int_field(sira_no, 'Sira no', default=1, min_val=1, max_val=50)
        konum_val, konum_error = parse_int_field(konum_no, 'Konum no', default=1, min_val=1, max_val=50)
        kat_val, kat_error = parse_int_field(kat_sayisi, 'Kat sayisi', default=1, min_val=1, max_val=10)
        cerceve_val, cerceve_error = parse_int_field(cerceve_sayisi, 'Cerceve sayisi',
                                                      default=10, min_val=1, max_val=50)
        ana_ari_val, ana_ari_error = parse_int_field(
            ana_ari_yili, 'Ana ari yili', default=None,
            min_val=2000, max_val=datetime.now().year + 1
        )
        errors = [status_error, sira_error, konum_error, kat_error, cerceve_error, ana_ari_error]
        if any(errors):
            flash_errors(errors)
            return render_template('fixed_hive_form.html', hive=None, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=False)

        conflict_error = fixed_hive_conflict(apiary_id, kovan_no, sira_val, konum_val)
        if conflict_error:
            flash(conflict_error, 'error')
            return render_template('fixed_hive_form.html', hive=None, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=False)

        # Fotograf
        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu = save_upload(file)
            if not fotograf_yolu:
                return render_template('fixed_hive_form.html', hive=None, apiary=apiary,
                                       statuses=FIXED_HIVE_STATUSES, edit=False)

        execute_db('''
            INSERT INTO fixed_hives (arilik_id, kovan_no, sira_no, konum_no, ana_ari_yili,
            kat_sayisi, cerceve_sayisi, durum, genel_not, fotograf_yolu)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (apiary_id, kovan_no, sira_val, konum_val, ana_ari_val,
              kat_val, cerceve_val, durum, genel_not or None, fotograf_yolu))

        flash('Sabit kovan basariyla eklendi.', 'success')
        return redirect(url_for('apiary_detail', id=apiary_id))

    return render_template('fixed_hive_form.html', hive=None, apiary=apiary,
                           statuses=FIXED_HIVE_STATUSES, edit=False)


@app.route('/fixed-hives/<int:id>')
def fixed_hive_detail(id):
    """Sabit kovan detay sayfasi."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi, a.id as arilik_id_val,
               a.latitude as apiary_latitude, a.longitude as apiary_longitude
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        if not session.get('authenticated'):
            flash('Kovan kaydı bulunamadı.', 'error')
            return redirect(url_for('public_home'))
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))
    if not session.get('authenticated'):
        source_context = get_message_source_context('fixed', id)
        return render_template(
            'public_hive.html',
            hive_context=public_hive_context(source_context),
        )

    # Son 5 kontrol kaydi
    inspections = query_db('''
        SELECT * FROM inspections WHERE kovan_id = ?
        ORDER BY kontrol_tarihi DESC LIMIT 5
    ''', (id,))

    return render_template('fixed_hive_detail.html', hive=hive, inspections=inspections)


@app.route('/fixed-hives/<int:id>/edit', methods=['GET', 'POST'])
def fixed_hive_edit(id):
    """Sabit kovan duzenleme."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    apiary = query_db('SELECT * FROM apiaries WHERE id = ?', (hive['arilik_id'],), one=True)

    if request.method == 'POST':
        kovan_no = request.form.get('kovan_no', '').strip()
        sira_no = request.form.get('sira_no', '1').strip()
        konum_no = request.form.get('konum_no', '1').strip()
        ana_ari_yili = request.form.get('ana_ari_yili', '').strip()
        kat_sayisi = request.form.get('kat_sayisi', '1').strip()
        cerceve_sayisi = request.form.get('cerceve_sayisi', '10').strip()
        durum = request.form.get('durum', 'Orta')
        genel_not = request.form.get('genel_not', '').strip()
        aktif = 1 if request.form.get('aktif') else 0

        if not kovan_no:
            flash('Kovan numarasi zorunludur.', 'error')
            return render_template('fixed_hive_form.html', hive=hive, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=True)

        durum, status_error = validate_choice(durum, FIXED_HIVE_STATUSES, 'Durum')
        sira_val, sira_error = parse_int_field(sira_no, 'Sira no', default=1, min_val=1, max_val=50)
        konum_val, konum_error = parse_int_field(konum_no, 'Konum no', default=1, min_val=1, max_val=50)
        kat_val, kat_error = parse_int_field(kat_sayisi, 'Kat sayisi', default=1, min_val=1, max_val=10)
        cerceve_val, cerceve_error = parse_int_field(cerceve_sayisi, 'Cerceve sayisi',
                                                      default=10, min_val=1, max_val=50)
        ana_ari_val, ana_ari_error = parse_int_field(
            ana_ari_yili, 'Ana ari yili', default=None,
            min_val=2000, max_val=datetime.now().year + 1
        )
        errors = [status_error, sira_error, konum_error, kat_error, cerceve_error, ana_ari_error]
        if any(errors):
            flash_errors(errors)
            return render_template('fixed_hive_form.html', hive=hive, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=True)

        conflict_error = fixed_hive_conflict(hive['arilik_id'], kovan_no, sira_val, konum_val,
                                             exclude_id=id)
        if conflict_error:
            flash(conflict_error, 'error')
            return render_template('fixed_hive_form.html', hive=hive, apiary=apiary,
                                   statuses=FIXED_HIVE_STATUSES, edit=True)

        fotograf_yolu = hive['fotograf_yolu']
        file = request.files.get('fotograf')
        if file and file.filename:
            new_foto = save_upload(file)
            if not new_foto:
                return render_template('fixed_hive_form.html', hive=hive, apiary=apiary,
                                       statuses=FIXED_HIVE_STATUSES, edit=True)
            fotograf_yolu = new_foto

        execute_db('''
            UPDATE fixed_hives SET kovan_no=?, sira_no=?, konum_no=?, ana_ari_yili=?,
            kat_sayisi=?, cerceve_sayisi=?, durum=?, genel_not=?, fotograf_yolu=?,
            aktif=?, guncelleme_tarihi=datetime('now','localtime')
            WHERE id=?
        ''', (kovan_no, sira_val, konum_val, ana_ari_val,
              kat_val, cerceve_val, durum, genel_not or None,
              fotograf_yolu, aktif, id))

        flash('Kovan guncellendi.', 'success')
        return redirect(url_for('fixed_hive_detail', id=id))

    return render_template('fixed_hive_form.html', hive=hive, apiary=apiary,
                           statuses=FIXED_HIVE_STATUSES, edit=True)


@app.route('/fixed-hives/<int:id>/qr')
def fixed_hive_qr(id):
    """Sabit kovan QR kod sayfasi."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    qr_path = generate_qr_code(id)
    if hive['qr_kod_yolu'] != qr_path:
        execute_db('UPDATE fixed_hives SET qr_kod_yolu=? WHERE id=?', (qr_path, id))

    return render_template(
        'qr_view.html',
        hive=hive,
        qr_path=qr_path,
        qr_label=hive['kovan_no'],
        qr_sublabel=hive['arilik_adi'],
        detail_url=url_for('fixed_hive_detail', id=id),
        public_qr_url=build_public_url('public_qr_fixed', id=id),
        back_label=f"{hive['kovan_no']} - {hive['arilik_adi']}",
        download_name=f"qr_{hive['kovan_no']}.png",
    )


# ---------------------------------------------------------------------------
# Modul 3: Kontrol Kayitlari
# ---------------------------------------------------------------------------

@app.route('/fixed-hives/<int:id>/inspections/new', methods=['GET', 'POST'])
def inspection_new(id):
    """Yeni kontrol kaydi ekleme."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    if request.method == 'POST':
        kontrol_tarihi = request.form.get('kontrol_tarihi', '').strip()
        if not kontrol_tarihi:
            kontrol_tarihi = datetime.now().strftime('%Y-%m-%d')
        kontrol_tarihi, date_error = validate_date(kontrol_tarihi, 'Kontrol tarihi', required=True)

        fields = {}
        errors = [date_error]
        for key in INSPECTION_CHOICES:
            value, choice_error = validate_choice(
                request.form.get(key, INSPECTION_CHOICES[key][0]),
                INSPECTION_CHOICES[key],
                key
            )
            fields[key] = value
            errors.append(choice_error)

        if any(errors):
            flash_errors(errors)
            return render_template('inspection_form.html', hive=hive, choices=INSPECTION_CHOICES)

        notlar = request.form.get('notlar', '').strip()

        # Fotograf
        fotograf_yolu = None
        file = request.files.get('fotograf')
        if file and file.filename:
            fotograf_yolu = save_upload(file)
            if not fotograf_yolu:
                return render_template('inspection_form.html', hive=hive, choices=INSPECTION_CHOICES)

        execute_db('''
            INSERT INTO inspections (kovan_id, kontrol_tarihi, ari_yogunlugu, yavru_durumu,
            bal_durumu, polen_gelisi, ana_ari_goruldu, yumurta_var, hastalik_belirtisi,
            saldirganlik, besleme_yapildi, ilaclama_yapildi, notlar, fotograf_yolu)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (id, kontrol_tarihi, fields['ari_yogunlugu'], fields['yavru_durumu'],
              fields['bal_durumu'], fields['polen_gelisi'], fields['ana_ari_goruldu'],
              fields['yumurta_var'], fields['hastalik_belirtisi'], fields['saldirganlik'],
              fields['besleme_yapildi'], fields['ilaclama_yapildi'],
              notlar or None, fotograf_yolu))

        # Kovanin son kontrol tarihini guncelle
        execute_db('''
            UPDATE fixed_hives SET son_kontrol_tarihi=?,
            guncelleme_tarihi=datetime('now','localtime')
            WHERE id=?
        ''', (kontrol_tarihi, id))

        flash('Kontrol kaydi eklendi.', 'success')
        return redirect(url_for('fixed_hive_detail', id=id))

    return render_template('inspection_form.html', hive=hive, choices=INSPECTION_CHOICES)


@app.route('/fixed-hives/<int:id>/inspections')
def inspection_history(id):
    """Kontrol gecmisi."""
    hive = query_db('''
        SELECT fh.*, a.arilik_adi
        FROM fixed_hives fh
        JOIN apiaries a ON fh.arilik_id = a.id
        WHERE fh.id = ?
    ''', (id,), one=True)
    if not hive:
        flash('Kovan bulunamadi.', 'error')
        return redirect(url_for('apiary_list'))

    inspections = query_db('''
        SELECT * FROM inspections WHERE kovan_id = ?
        ORDER BY kontrol_tarihi DESC
    ''', (id,))

    return render_template('inspection_history.html', hive=hive, inspections=inspections)


# ---------------------------------------------------------------------------
# Calistirma
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    if not os.path.exists(DB_PATH):
        print("[UYARI] Veritabani bulunamadi. Lutfen once 'python init_db.py' calistirin.")
    host = os.environ.get('ALI_BABA_HOST', '127.0.0.1')
    port = int(os.environ.get('ALI_BABA_PORT', '5000'))
    debug = os.environ.get('ALI_BABA_DEBUG', '0') == '1'
    app.run(host=host, port=port, debug=debug)
