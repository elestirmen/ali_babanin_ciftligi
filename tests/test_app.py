from io import BytesIO
import os
import shutil
import sqlite3
import tempfile
import unittest
import zipfile
from pathlib import Path

from PIL import Image
from werkzeug.datastructures import FileStorage
from openpyxl import load_workbook


_tmp = tempfile.TemporaryDirectory()
_base = Path(_tmp.name)
os.environ['ALI_BABA_DB_PATH'] = str(_base / 'test.db')
os.environ['ALI_BABA_UPLOAD_FOLDER'] = str(_base / 'uploads')
os.environ['ALI_BABA_QR_FOLDER'] = str(_base / 'qrcodes')
os.environ['ALI_BABA_PUBLIC_UPLOAD_FOLDER'] = str(_base / 'public_uploads')
os.environ['ALI_BABA_BACKUP_FOLDER'] = str(_base / 'backups')
os.environ['ALI_BABA_SECRET_KEY'] = 'test-secret'
os.environ['ALI_BABA_PASSWORD'] = 'alibaba'

import app as app_module  # noqa: E402
import init_db  # noqa: E402


class AppSecurityAndValidationTests(unittest.TestCase):
    @classmethod
    def tearDownClass(cls):
        _tmp.cleanup()

    def setUp(self):
        db_path = Path(app_module.DB_PATH)
        if db_path.exists():
            db_path.unlink()
        Path(app_module.UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
        Path(app_module.QR_FOLDER).mkdir(parents=True, exist_ok=True)
        Path(app_module.PUBLIC_UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
        backup_path = Path(app_module.BACKUP_FOLDER)
        if backup_path.exists():
            shutil.rmtree(backup_path)
        backup_path.mkdir(parents=True, exist_ok=True)
        init_db.DB_PATH = app_module.DB_PATH
        init_db.create_tables()
        app_module.app.config.update(TESTING=True)
        with app_module.RATE_LIMIT_LOCK:
            app_module.RATE_LIMIT_BUCKETS.clear()
        self.client = app_module.app.test_client()

    def login(self):
        response = self.client.post('/login', data={'password': 'alibaba'})
        self.assertEqual(response.status_code, 302)
        with self.client.session_transaction() as sess:
            return sess['_csrf_token']

    def count_rows(self, table):
        conn = sqlite3.connect(app_module.DB_PATH)
        try:
            return conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
        finally:
            conn.close()

    def fetch_one(self, query, args=()):
        conn = sqlite3.connect(app_module.DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            return conn.execute(query, args).fetchone()
        finally:
            conn.close()

    def test_public_home_is_visible_and_admin_requires_login(self):
        response = self.client.get('/')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b"Ali Baba", response.data)

        response = self.client.get('/admin')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.headers['Location'])

    def test_login_accepts_alibaba_password(self):
        response = self.client.post('/login', data={'password': 'yanlis'})
        self.assertEqual(response.status_code, 200)

        response = self.client.post('/login', data={'password': 'alibaba'})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers['Location'], '/admin')

    def test_security_headers_and_session_cookie_flags(self):
        response = self.client.get('/')
        self.assertEqual(response.headers['X-Content-Type-Options'], 'nosniff')
        self.assertEqual(response.headers['X-Frame-Options'], 'SAMEORIGIN')
        self.assertEqual(response.headers['Referrer-Policy'], 'strict-origin-when-cross-origin')
        self.assertIn("default-src 'self'", response.headers['Content-Security-Policy'])

        response = self.client.post('/login', data={'password': 'alibaba'})
        cookie = response.headers.get('Set-Cookie', '')
        self.assertIn('HttpOnly', cookie)
        self.assertIn('SameSite=Lax', cookie)

    def test_failed_login_rate_limit(self):
        old_limit = app_module.app.config['LOGIN_FAILED_RATE_LIMIT']
        old_window = app_module.app.config['LOGIN_FAILED_RATE_WINDOW']
        app_module.app.config['LOGIN_FAILED_RATE_LIMIT'] = 2
        app_module.app.config['LOGIN_FAILED_RATE_WINDOW'] = 600
        try:
            for _ in range(2):
                response = self.client.post('/login', data={'password': 'yanlis'})
                self.assertEqual(response.status_code, 200)

            response = self.client.post('/login', data={'password': 'yanlis'})
            self.assertEqual(response.status_code, 429)
        finally:
            app_module.app.config['LOGIN_FAILED_RATE_LIMIT'] = old_limit
            app_module.app.config['LOGIN_FAILED_RATE_WINDOW'] = old_window
            with app_module.RATE_LIMIT_LOCK:
                app_module.RATE_LIMIT_BUCKETS.clear()

    def test_public_message_can_be_left_without_login(self):
        response = self.client.post('/mesaj-birak', data={
            'ad': 'Ziyaretci',
            'iletisim': 'ziyaretci@example.com',
            'konu': 'Bal',
            'mesaj': 'Merhaba, bal hakkında bilgi almak istiyorum.',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers['Location'], '/mesaj-birak')
        self.assertEqual(self.count_rows('public_messages'), 1)
        message = self.fetch_one('SELECT kategori FROM public_messages')
        self.assertEqual(message['kategori'], 'Diğer')

    def test_public_message_rate_limit(self):
        old_limit = app_module.app.config['PUBLIC_MESSAGE_RATE_LIMIT']
        old_window = app_module.app.config['PUBLIC_MESSAGE_RATE_WINDOW']
        app_module.app.config['PUBLIC_MESSAGE_RATE_LIMIT'] = 2
        app_module.app.config['PUBLIC_MESSAGE_RATE_WINDOW'] = 600
        ip = '203.0.113.25'
        try:
            for index in range(2):
                response = self.client.post('/mesaj-birak', data={
                    'ad': 'Ziyaretci',
                    'mesaj': f'Mesaj {index}',
                }, environ_overrides={'REMOTE_ADDR': ip})
                self.assertEqual(response.status_code, 302)

            response = self.client.post('/mesaj-birak', data={
                'ad': 'Ziyaretci',
                'mesaj': 'Ucuncu mesaj',
            }, environ_overrides={'REMOTE_ADDR': ip})
            self.assertEqual(response.status_code, 429)
            self.assertEqual(self.count_rows('public_messages'), 2)
        finally:
            app_module.app.config['PUBLIC_MESSAGE_RATE_LIMIT'] = old_limit
            app_module.app.config['PUBLIC_MESSAGE_RATE_WINDOW'] = old_window
            with app_module.RATE_LIMIT_LOCK:
                app_module.RATE_LIMIT_BUCKETS.clear()

    def test_public_hive_qr_landing_and_context_message(self):
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Gizli Arilik', 38.63, 34.82)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, sira_no, konum_no, durum)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'K7', 1, 1, 'Orta')
        )

        response = self.client.get(f'/fixed-hives/{fixed_id}')
        self.assertEqual(response.status_code, 200)
        self.assertIn('Doğanın küçük çalışanlarından'.encode('utf-8'), response.data)
        self.assertNotIn('Kovanın konumu'.encode('utf-8'), response.data)
        self.assertNotIn(b'publicHiveMap', response.data)
        self.assertNotIn(b'38.63', response.data)
        self.assertNotIn(b'34.82', response.data)
        self.assertNotIn('Bakım notları'.encode('utf-8'), response.data)
        self.assertIn('Sorun Bildir'.encode('utf-8'), response.data)
        self.assertNotIn('Yeni Kontrol Kaydı'.encode('utf-8'), response.data)

        response = self.client.get(f'/q/fixed/{fixed_id}')
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers['Location'], f'/fixed-hives/{fixed_id}')

        response = self.client.post('/mesaj-birak', data={
            'kaynak_turu': 'fixed',
            'kaynak_id': str(fixed_id),
            'kategori': 'Kovan devrilmiş / zarar görmüş',
            'konu': 'Kovan etiketi',
            'mesaj': 'Kovan yan yatmış görünüyor.',
        })
        self.assertEqual(response.status_code, 302)
        message = self.fetch_one('''
            SELECT ad, kategori, kaynak_turu, kaynak_id
            FROM public_messages
        ''')
        self.assertEqual(message['ad'], 'İsimsiz ziyaretçi')
        self.assertEqual(message['kategori'], 'Kovan devrilmiş / zarar görmüş')
        self.assertEqual(message['kaynak_turu'], 'fixed')
        self.assertEqual(message['kaynak_id'], fixed_id)

        self.login()
        response = self.client.get('/admin/messages?q=K7')
        self.assertEqual(response.status_code, 200)
        self.assertIn('Sabit kovan K7'.encode('utf-8'), response.data)
        self.assertIn('Kovan yan yatmış'.encode('utf-8'), response.data)

        response = self.client.get(f'/fixed-hives/{fixed_id}')
        self.assertEqual(response.status_code, 200)
        self.assertIn('Yeni Kontrol Kaydı'.encode('utf-8'), response.data)

    def test_admin_message_filters_and_status_update(self):
        token = self.login()
        app_module.execute_db('''
            INSERT INTO public_messages (ad, iletisim, kategori, konu, mesaj)
            VALUES (?, ?, ?, ?, ?)
        ''', ('Ziyaretci', 'ziyaretci@example.com', 'Ziyaret talebi', 'Ziyaret', 'Hafta sonu gelebilir miyiz?'))

        response = self.client.get('/admin/messages?durum=Yeni&kategori=Ziyaret+talebi&q=hafta')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Ziyaretci', response.data)

        response = self.client.post('/admin/messages/1/status', data={
            'csrf_token': token,
            'durum': 'Ziyaret planlandı',
            'yanit_notu': 'Cumartesi uygun.',
        })
        self.assertEqual(response.status_code, 302)
        message = self.fetch_one('SELECT durum, yanit_notu, okundu FROM public_messages WHERE id = 1')
        self.assertEqual(message['durum'], 'Ziyaret planlandı')
        self.assertEqual(message['yanit_notu'], 'Cumartesi uygun.')
        self.assertEqual(message['okundu'], 1)

    def test_admin_messages_require_login(self):
        response = self.client.get('/admin/messages')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.headers['Location'])

    def test_csrf_required_for_post_routes(self):
        self.login()
        response = self.client.post('/apiaries/new', data={'arilik_adi': 'Test'})
        self.assertEqual(response.status_code, 400)

    def test_location_picker_is_rendered_on_coordinate_forms(self):
        self.login()

        response = self.client.get('/swarm-hives/new')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'locationPickerMap', response.data)
        self.assertIn(b'js/location_picker.js', response.data)

        response = self.client.get('/apiaries/new')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'locationPickerMap', response.data)
        self.assertIn(b'js/location_picker.js', response.data)

    def test_invalid_swarm_status_is_rejected(self):
        token = self.login()
        response = self.client.post('/swarm-hives/new', data={
            'csrf_token': token,
            'ad': 'Test Kovan',
            'durum': 'Gecersiz Durum',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.count_rows('swarm_hives'), 0)

    def test_duplicate_fixed_hive_number_is_rejected(self):
        token = self.login()
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Ana Arilik', 38.63, 34.82)
        )
        app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, sira_no, konum_no, durum)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'K1', 1, 1, 'Orta')
        )

        response = self.client.post(f'/apiaries/{apiary_id}/fixed-hives/new', data={
            'csrf_token': token,
            'kovan_no': 'K1',
            'sira_no': '2',
            'konum_no': '2',
            'kat_sayisi': '1',
            'cerceve_sayisi': '10',
            'durum': 'Orta',
        })
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.count_rows('fixed_hives'), 1)

    def test_public_url_overrides_absolute_url_generation(self):
        previous_public_url = app_module.PUBLIC_URL
        app_module.PUBLIC_URL = 'https://alibaba.urgup.keenetic.link'
        try:
            with app_module.app.test_request_context('/', base_url='http://internal.local'):
                url = app_module.build_public_url('fixed_hive_detail', id=7)
        finally:
            app_module.PUBLIC_URL = previous_public_url

        self.assertEqual(url, 'https://alibaba.urgup.keenetic.link/fixed-hives/7')

    def test_private_media_requires_login_and_public_media_is_open(self):
        private_file = Path(app_module.UPLOAD_FOLDER) / 'private.txt'
        public_file = Path(app_module.PUBLIC_UPLOAD_FOLDER) / 'public.txt'
        private_file.write_text('private', encoding='utf-8')
        public_file.write_text('public', encoding='utf-8')

        response = self.client.get('/public-media/public.txt')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b'public')
        response.close()

        response = self.client.get('/media/uploads/private.txt')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.headers['Location'])

        self.login()
        response = self.client.get('/media/uploads/private.txt')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.data, b'private')
        response.close()

    def test_invalid_public_post_upload_does_not_create_record(self):
        token = self.login()
        response = self.client.post('/admin/posts/new', data={
            'csrf_token': token,
            'baslik': 'Gecersiz fotograf',
            'kategori': 'Duyuru',
            'icerik': 'Bu kayit olusmamali.',
            'yayin_tarihi': '2026-04-26',
            'yayinla': 'on',
            'fotograf': (BytesIO(b'not an image'), 'bad.jpg'),
        })

        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.count_rows('public_posts'), 0)

    def test_saved_uploads_are_rewritten_without_exif(self):
        image = Image.new('RGB', (4, 4), color='white')
        exif = Image.Exif()
        exif[271] = 'Test Camera'
        buffer = BytesIO()
        image.save(buffer, format='JPEG', exif=exif.tobytes())
        buffer.seek(0)

        with app_module.app.test_request_context('/'):
            path = app_module.save_upload(FileStorage(stream=buffer, filename='photo.jpg'))

        self.assertIsNotNone(path)
        saved_path = app_module.media_abs_path(path)
        with Image.open(saved_path) as saved:
            self.assertEqual(dict(saved.getexif()), {})

    def test_upload_rejects_images_over_pixel_limit(self):
        old_limit = app_module.app.config['MAX_IMAGE_PIXELS']
        old_pillow_limit = Image.MAX_IMAGE_PIXELS
        app_module.app.config['MAX_IMAGE_PIXELS'] = 10
        try:
            image = Image.new('RGB', (4, 4), color='white')
            buffer = BytesIO()
            image.save(buffer, format='PNG')
            buffer.seek(0)

            with app_module.app.test_request_context('/'):
                path = app_module.save_upload(FileStorage(stream=buffer, filename='too-large.png'))

            self.assertIsNone(path)
        finally:
            app_module.app.config['MAX_IMAGE_PIXELS'] = old_limit
            Image.MAX_IMAGE_PIXELS = old_pillow_limit

    def test_admin_hives_table_and_quick_updates(self):
        token = self.login()
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi) VALUES (?)',
            ('Ana Arilik',)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, sira_no, konum_no, durum, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (apiary_id, 'K1', 1, 1, 'Orta', '2026-04-20', 1)
        )
        swarm_id = app_module.execute_db(
            '''INSERT INTO swarm_hives
               (ad, durum, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?)''',
            ('Dere Kenari', 'Boş', '2026-04-20', 1)
        )

        response = self.client.get('/admin/hives')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'Dere Kenari', response.data)
        self.assertIn(b'K1', response.data)

        response = self.client.post(f'/admin/hives/fixed/{fixed_id}/quick-update', data={
            'csrf_token': token,
            'durum': 'Güçlü',
            'son_kontrol_tarihi': '2026-04-26',
            'aktif': 'on',
        })
        self.assertEqual(response.status_code, 302)
        fixed = self.fetch_one('SELECT durum, son_kontrol_tarihi, aktif FROM fixed_hives WHERE id = ?', (fixed_id,))
        self.assertEqual(fixed['durum'], 'Güçlü')
        self.assertEqual(fixed['son_kontrol_tarihi'], '2026-04-26')
        self.assertEqual(fixed['aktif'], 1)

        response = self.client.post(f'/admin/hives/swarm/{swarm_id}/quick-update', data={
            'csrf_token': token,
            'durum': 'Oğul girdi',
            'son_kontrol_tarihi': '2026-04-26',
        })
        self.assertEqual(response.status_code, 302)
        swarm = self.fetch_one('SELECT durum, son_kontrol_tarihi, aktif FROM swarm_hives WHERE id = ?', (swarm_id,))
        self.assertEqual(swarm['durum'], 'Oğul girdi')
        self.assertEqual(swarm['son_kontrol_tarihi'], '2026-04-26')
        self.assertEqual(swarm['aktif'], 0)

    def test_delete_routes_require_confirmation_and_cascade(self):
        token = self.login()
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi) VALUES (?)',
            ('Silinecek Arilik',)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, sira_no, konum_no, durum)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'S1', 1, 1, 'Orta')
        )
        app_module.execute_db(
            '''INSERT INTO inspections
               (kovan_id, kontrol_tarihi, ari_yogunlugu)
               VALUES (?, ?, ?)''',
            (fixed_id, '2026-04-26', 'Orta')
        )
        swarm_id = app_module.execute_db(
            '''INSERT INTO swarm_hives (ad, durum)
               VALUES (?, ?)''',
            ('Silinecek Ogul', 'Boş')
        )
        app_module.execute_db(
            '''INSERT INTO swarm_inspections
               (swarm_hive_id, kontrol_tarihi, durum)
               VALUES (?, ?, ?)''',
            (swarm_id, '2026-04-26', 'Boş')
        )

        response = self.client.post(f'/swarm-hives/{swarm_id}/delete', data={
            'csrf_token': token,
            'confirm_text': 'yanlis',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.count_rows('swarm_hives'), 1)

        response = self.client.post(f'/swarm-hives/{swarm_id}/delete', data={
            'csrf_token': token,
            'confirm_text': 'SIL',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.count_rows('swarm_hives'), 0)
        self.assertEqual(self.count_rows('swarm_inspections'), 0)

        response = self.client.post(f'/fixed-hives/{fixed_id}/delete', data={
            'csrf_token': token,
            'confirm_text': 'SIL',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.count_rows('fixed_hives'), 0)
        self.assertEqual(self.count_rows('inspections'), 0)

        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, sira_no, konum_no, durum)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'S2', 1, 1, 'Orta')
        )
        app_module.execute_db(
            '''INSERT INTO inspections
               (kovan_id, kontrol_tarihi, ari_yogunlugu)
               VALUES (?, ?, ?)''',
            (fixed_id, '2026-04-26', 'Orta')
        )

        response = self.client.post(f'/apiaries/{apiary_id}/delete', data={
            'csrf_token': token,
            'confirm_text': 'SIL',
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(self.count_rows('apiaries'), 0)
        self.assertEqual(self.count_rows('fixed_hives'), 0)
        self.assertEqual(self.count_rows('inspections'), 0)

    def test_swarm_qr_inspection_history_and_map_focus(self):
        token = self.login()
        swarm_id = app_module.execute_db(
            '''INSERT INTO swarm_hives
               (ad, latitude, longitude, durum, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?, ?, ?)''',
            ('Dere Kenarı', 38.63, 34.82, 'Boş', '2026-04-20', 1)
        )

        response = self.client.get(f'/swarm-hives/{swarm_id}')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'/admin?focus=swarm-', response.data)
        self.assertIn('Yeni Kontrol Kaydı'.encode('utf-8'), response.data)
        self.assertIn('QR Kod'.encode('utf-8'), response.data)

        response = self.client.get(f'/swarm-hives/{swarm_id}/qr')
        self.assertEqual(response.status_code, 200)
        qr_row = self.fetch_one('SELECT qr_kod_yolu FROM swarm_hives WHERE id = ?', (swarm_id,))
        self.assertTrue(qr_row['qr_kod_yolu'].startswith('qrcodes/qr_swarm_'))
        self.assertTrue(Path(app_module.media_abs_path(qr_row['qr_kod_yolu'])).exists())

        response = self.client.post(f'/swarm-hives/{swarm_id}/inspections/new', data={
            'csrf_token': token,
            'kontrol_tarihi': '2026-04-26',
            'durum': 'Arı hareketi var',
            'ari_gelis_tarihi': '2026-04-25',
            'notlar': 'Sabah arı hareketi görüldü.',
        })
        self.assertEqual(response.status_code, 302)
        swarm = self.fetch_one(
            'SELECT durum, son_kontrol_tarihi FROM swarm_hives WHERE id = ?',
            (swarm_id,)
        )
        self.assertEqual(swarm['durum'], 'Arı hareketi var')
        self.assertEqual(swarm['son_kontrol_tarihi'], '2026-04-26')
        inspection = self.fetch_one(
            '''SELECT durum, ari_gelis_tarihi, notlar
               FROM swarm_inspections WHERE swarm_hive_id = ?''',
            (swarm_id,)
        )
        self.assertEqual(inspection['ari_gelis_tarihi'], '2026-04-25')
        self.assertIn('Sabah arı hareketi', inspection['notlar'])

        response = self.client.get(f'/swarm-hives/{swarm_id}/inspections')
        self.assertEqual(response.status_code, 200)
        self.assertIn('Arı hareketi var'.encode('utf-8'), response.data)
        self.assertIn('Sabah arı hareketi'.encode('utf-8'), response.data)

        response = self.client.get(f'/api/map-data?focus=swarm-{swarm_id}')
        self.assertEqual(response.status_code, 200)
        data = response.get_json()
        focused = [item for item in data['swarm_hives'] if item['id'] == swarm_id]
        self.assertEqual(len(focused), 1)
        self.assertTrue(focused[0]['focused'])

    def test_fixed_hive_map_focus_includes_normal_hive(self):
        self.login()
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Ana Arılık', 38.64, 34.83)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, durum, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'K1', 'Güçlü', '2099-01-01', 1)
        )

        response = self.client.get('/api/map-data')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['fixed_hives'], [])

        response = self.client.get(f'/api/map-data?focus=fixed-{fixed_id}')
        self.assertEqual(response.status_code, 200)
        fixed_hives = response.get_json()['fixed_hives']
        self.assertEqual(len(fixed_hives), 1)
        self.assertEqual(fixed_hives[0]['id'], fixed_id)
        self.assertTrue(fixed_hives[0]['focused'])

        response = self.client.get(f'/fixed-hives/{fixed_id}')
        self.assertEqual(response.status_code, 200)
        self.assertIn(f'/admin?focus=fixed-{fixed_id}'.encode('utf-8'), response.data)

    def test_admin_backups_require_login(self):
        response = self.client.get('/admin/backups')
        self.assertEqual(response.status_code, 302)
        self.assertIn('/login', response.headers['Location'])

    def test_export_routes_require_login(self):
        routes = [
            '/admin/export',
            '/admin/export/swarm-hives.csv',
            '/admin/export/swarm-hives.xlsx',
            '/admin/export/fixed-hives.csv',
            '/admin/export/fixed-hives.xlsx',
            '/admin/export/inspections.csv',
            '/admin/export/inspections.xlsx',
            '/admin/export/apiary-summary.csv',
            '/admin/export/apiary-summary.xlsx',
            '/admin/export/all.xlsx',
        ]
        for route in routes:
            with self.subTest(route=route):
                response = self.client.get(route)
                self.assertEqual(response.status_code, 302)
                self.assertIn('/login', response.headers['Location'])

    def test_csv_export_routes_return_downloads_after_login(self):
        self.login()
        app_module.execute_db(
            '''INSERT INTO swarm_hives
               (ad, latitude, longitude, durum, kurulum_tarihi, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            ('Dere Kenarı', 38.63, 34.82, 'Boş', '2026-04-01', '2026-04-20', 1)
        )
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Ana Arılık', 38.64, 34.83)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, durum, son_kontrol_tarihi, aktif)
               VALUES (?, ?, ?, ?, ?)''',
            (apiary_id, 'K1', 'Güçlü', '2026-04-21', 1)
        )
        app_module.execute_db(
            '''INSERT INTO inspections (kovan_id, kontrol_tarihi, ari_yogunlugu, notlar)
               VALUES (?, ?, ?, ?)''',
            (fixed_id, '2026-04-22', 'Güçlü', 'Türkçe not')
        )

        expected = {
            '/admin/export/swarm-hives.csv': 'ogul_kovanlari_',
            '/admin/export/fixed-hives.csv': 'sabit_kovanlar_',
            '/admin/export/inspections.csv': 'kontrol_kayitlari_',
            '/admin/export/apiary-summary.csv': 'arilik_ozeti_',
        }
        for route, prefix in expected.items():
            with self.subTest(route=route):
                response = self.client.get(route)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(response.content_type, 'text/csv; charset=utf-8')
                disposition = response.headers.get('Content-Disposition', '')
                self.assertIn('attachment', disposition)
                self.assertIn(prefix, disposition)
                self.assertTrue(response.data.startswith(b'\xef\xbb\xbf'))

    def test_export_page_lists_csv_and_excel_downloads_in_table(self):
        self.login()
        response = self.client.get('/admin/export')
        self.assertEqual(response.status_code, 200)
        self.assertIn(b'<table', response.data)
        self.assertIn(b'/admin/export/swarm-hives.csv', response.data)
        self.assertIn(b'/admin/export/swarm-hives.xlsx', response.data)
        self.assertIn(b'/admin/export/all.xlsx', response.data)
        self.assertIn('Excel indir'.encode('utf-8'), response.data)

    def test_single_sheet_excel_exports_return_downloads_after_login(self):
        self.login()
        app_module.execute_db(
            '''INSERT INTO swarm_hives
               (ad, latitude, longitude, durum)
               VALUES (?, ?, ?, ?)''',
            ('Dere Kenarı', 38.63, 34.82, 'Boş')
        )
        apiary_id = app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Ana Arılık', 38.64, 34.83)
        )
        fixed_id = app_module.execute_db(
            '''INSERT INTO fixed_hives
               (arilik_id, kovan_no, durum)
               VALUES (?, ?, ?)''',
            (apiary_id, 'K1', 'Güçlü')
        )
        app_module.execute_db(
            '''INSERT INTO inspections (kovan_id, kontrol_tarihi, ari_yogunlugu)
               VALUES (?, ?, ?)''',
            (fixed_id, '2026-04-22', 'Güçlü')
        )

        expected = {
            '/admin/export/swarm-hives.xlsx': ('ogul_kovanlari_', 'Oğul Kovanları'),
            '/admin/export/fixed-hives.xlsx': ('sabit_kovanlar_', 'Sabit Kovanlar'),
            '/admin/export/inspections.xlsx': ('kontrol_kayitlari_', 'Kontrol Kayıtları'),
            '/admin/export/apiary-summary.xlsx': ('arilik_ozeti_', 'Arılık Özeti'),
        }
        for route, (prefix, sheet_name) in expected.items():
            with self.subTest(route=route):
                response = self.client.get(route)
                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response.mimetype,
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                disposition = response.headers.get('Content-Disposition', '')
                self.assertIn('attachment', disposition)
                self.assertIn(prefix, disposition)
                workbook = load_workbook(BytesIO(response.data))
                self.assertEqual(workbook.sheetnames, [sheet_name])
                sheet = workbook[sheet_name]
                self.assertGreater(len(sheet.tables), 0)
                if sheet_name in {'Oğul Kovanları', 'Sabit Kovanlar', 'Arılık Özeti'}:
                    headers = [cell.value for cell in sheet[1]]
                    maps_col = headers.index('Google Maps yol tarifi linki') + 1
                    self.assertIsNotNone(sheet.cell(row=2, column=maps_col).hyperlink)

    def test_excel_export_returns_expected_workbook_after_login(self):
        self.login()
        app_module.execute_db(
            'INSERT INTO apiaries (arilik_adi, latitude, longitude) VALUES (?, ?, ?)',
            ('Ana Arılık', 38.64, 34.83)
        )

        response = self.client.get('/admin/export/all.xlsx')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.mimetype,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        disposition = response.headers.get('Content-Disposition', '')
        self.assertIn('attachment', disposition)
        self.assertIn('ali_baba_tum_veriler_', disposition)

        workbook = load_workbook(BytesIO(response.data))
        self.assertEqual(workbook.sheetnames, [
            'Oğul Kovanları',
            'Arılıklar',
            'Sabit Kovanlar',
            'Kontrol Kayıtları',
            'Arılık Özeti',
        ])
        for sheet in workbook.worksheets:
            self.assertGreater(len(sheet.tables), 0)

        apiary_sheet = workbook['Arılıklar']
        headers = [cell.value for cell in apiary_sheet[1]]
        maps_col = headers.index('Google Maps yol tarifi linki') + 1
        self.assertIsNotNone(apiary_sheet.cell(row=2, column=maps_col).hyperlink)

    def test_backup_create_adds_zip_with_database(self):
        token = self.login()
        app_module.execute_db(
            '''INSERT INTO swarm_hives (ad, durum)
               VALUES (?, ?)''',
            ('Yedek Kovanı', 'Boş')
        )

        response = self.client.post('/admin/backups/create', data={'csrf_token': token})
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.headers['Location'], '/admin/backups')

        backups = list(Path(app_module.BACKUP_FOLDER).glob('ali_baba_backup_*.zip'))
        self.assertEqual(len(backups), 1)
        with zipfile.ZipFile(backups[0]) as archive:
            self.assertIn(Path(app_module.DB_PATH).name, archive.namelist())

    def test_backup_download_and_delete(self):
        token = self.login()
        self.client.post('/admin/backups/create', data={'csrf_token': token})
        backup = next(Path(app_module.BACKUP_FOLDER).glob('ali_baba_backup_*.zip'))

        response = self.client.get(f'/admin/backups/{backup.name}/download')
        self.assertEqual(response.status_code, 200)
        self.assertIn('attachment', response.headers.get('Content-Disposition', ''))
        response.close()

        response = self.client.post(f'/admin/backups/{backup.name}/delete', data={'csrf_token': token})
        self.assertEqual(response.status_code, 302)
        self.assertFalse(backup.exists())

    def test_backup_create_survives_missing_data_folders(self):
        token = self.login()
        for folder in [
            app_module.UPLOAD_FOLDER,
            app_module.QR_FOLDER,
            app_module.PUBLIC_UPLOAD_FOLDER,
        ]:
            shutil.rmtree(folder, ignore_errors=True)

        response = self.client.post('/admin/backups/create', data={'csrf_token': token})
        self.assertEqual(response.status_code, 302)
        backups = list(Path(app_module.BACKUP_FOLDER).glob('ali_baba_backup_*.zip'))
        self.assertEqual(len(backups), 1)
        with zipfile.ZipFile(backups[0]) as archive:
            self.assertIn(Path(app_module.DB_PATH).name, archive.namelist())


if __name__ == '__main__':
    unittest.main()
